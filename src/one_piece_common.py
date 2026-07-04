import os
import re
import json
import glob
import shutil
import urllib.request
import urllib.error
import sys
import atexit
from urllib.parse import urljoin
from html import unescape
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# ============================================================
# PATH PROGETTO
# ============================================================
# I file .py stanno in src/. Le cartelle dati/output stanno un livello sopra:
#   bkp/  json/  out/  src/  stg/
# Funziona sia lanciando:
#   python src/one_piece_collection_build.py
# sia entrando in src/ e lanciando:
#   python one_piece_collection_build.py

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

if os.path.basename(SCRIPT_DIR).lower() == "src":
    PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
else:
    PROJECT_ROOT = SCRIPT_DIR

JSON_DIR = os.path.join(PROJECT_ROOT, "json")
STG_DIR = os.path.join(PROJECT_ROOT, "stg")
OUT_DIR = os.path.join(PROJECT_ROOT, "out")
BKP_DIR = os.path.join(PROJECT_ROOT, "bkp")
LOG_DIR = os.path.join(PROJECT_ROOT, "logs")

OUTPUT_BASE = "one_piece_collection"
OUTPUT_XLSX = os.path.join(OUT_DIR, f"{OUTPUT_BASE}.xlsx")
OUTPUT_JSON = os.path.join(OUT_DIR, f"{OUTPUT_BASE}.json")

RAW_CSV = os.path.join(STG_DIR, "bandai_cards_raw.csv")
SETS_CSV = os.path.join(STG_DIR, "one_piece_sets.csv")
ERRORS_CSV = os.path.join(STG_DIR, "one_piece_errors.csv")
FINAL_STG_CSV = os.path.join(STG_DIR, "one_piece_collection_stg.csv")
CARDMARKET_MERGED_CSV = os.path.join(STG_DIR, "cardmarket_prices_merged.csv")
CARDMARKET_EXPANSIONS_CSV = os.path.join(STG_DIR, "cardmarket_expansions.csv")
CARDMARKET_UPDATED_CSV = os.path.join(STG_DIR, "cardmarket_prices_updated.csv")
UPDATE_REPORT_CSV = os.path.join(STG_DIR, "price_update_report.csv")
UPDATED_STG_CSV = os.path.join(STG_DIR, "one_piece_collection_updated_stg.csv")
VALUE_HISTORY_CSV = os.path.join(STG_DIR, "value_history.csv")
JP_OFFICIAL_RAW_CSV = os.path.join(STG_DIR, "bandai_cards_jp_raw.csv")
JP_RARITY_OVERRIDES_CSV = os.path.join(STG_DIR, "jp_rarity_overrides.csv")

PRICE_GUIDE_PATTERN = "price_guide*.json"
PRODUCTS_SINGLES_PATTERN = "products_singles*.json"
PRODUCTS_NONSINGLES_PATTERN = "products_nonsingles*.json"

CARDMARKET_FALLBACK_JSON_URLS = {
    "singles": "https://downloads.s3.cardmarket.com/productCatalog/productList/products_singles_18.json",
    "nonsingles": "https://downloads.s3.cardmarket.com/productCatalog/productList/products_nonsingles_18.json",
    "priceguide": "https://downloads.s3.cardmarket.com/productCatalog/priceGuide/price_guide_18.json",
}

CARDMARKET_DATA_PAGES = {
    "price_guide": "https://www.cardmarket.com/en/Magic/Data/Price-Guide",
    "product_list": "https://www.cardmarket.com/en/Magic/Data/Product-List",
}

CARDMARKET_JSON_SOURCES_JSON = os.path.join(STG_DIR, "cardmarket_json_sources.json")

AUTO_DOWNLOAD_CARDMARKET_JSONS = True

PRICE_SOURCE_COLUMN = "trend"
EXCLUDE_PROMOS = True

IGNORED_REPO_FILES = {".gitkeep", ".gitignore", "README.md"}

PROMO_KEYWORDS = [
    "promo", "promotion", "tournament", "winner", "event pack", "judge pack",
    "pre-release", "pre release", "regional", "standard battle", "dash pack",
    "demo deck", "premium card collection", "premium card", "gift collection",
    "double pack", "welcome pack", "anniversary", "championship", "store championship",
    "sealed battle", "participation pack", "top player", "finalist", "don!! card pack",
    "treasure boosters", "sound loader", "red envelope", "film red promotion",
    "box promotion", "revision pack", "release event", "standard pack", "v jump",
    "cs 2023", "cs 2024", "cs finalist"
]

KNOWN_EXPANSIONS = [
    ("romance dawn", "OP01"), ("paramount war", "OP02"),
    ("pillars of strength", "OP03"), ("kingdoms of intrigue", "OP04"),
    ("awakening of the new era", "OP05"), ("wings of the captain", "OP06"),
    ("500 years into the future", "OP07"), ("two legends", "OP08"),
    ("emperors in the new world", "OP09"), ("royal blood", "OP10"),
    ("a fist of divine speed", "OP11"), ("legacy of the master", "OP12"),
    ("memorial collection", "EB01"), ("the best", "PRB01"),
    ("anime 25th collection", "EB02"),
    ("straw hat crew", "ST01"), ("worst generation", "ST02"),
    ("seven warlords", "ST03"), ("animal kingdom pirates", "ST04"),
    ("film edition", "ST05"), ("absolute justice", "ST06"),
    ("big mom pirates", "ST07"), ("monkey.d.luffy", "ST08"),
    ("yamato", "ST09"), ("the three captains", "ST10"),
    ("uta", "ST11"), ("zoro & sanji", "ST12"),
    ("the three brothers", "ST13"), ("3d2y", "ST14"),
    ("edward.newgate", "ST15"), ("green uta", "ST16"),
    ("donquixote doflamingo", "ST17"), ("purple monkey.d.luffy", "ST18"),
    ("smoker", "ST19"), ("charlotte katakuri", "ST20"),
    ("ex gear 5", "ST21")
]




def clean_for_excel(value):
    """Rende sicuri i testi prima di scriverli in Excel.

    openpyxl non accetta alcuni caratteri di controllo e le celle hanno un limite
    pratico di circa 32k caratteri. Tenerla in common evita crash nelle funzioni
    centralizzate di Dashboard/Excel.
    """
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    return value[:32000] if len(value) > 32000 else value

def ensure_dirs():
    for d in [JSON_DIR, STG_DIR, OUT_DIR, BKP_DIR, LOG_DIR]:
        os.makedirs(d, exist_ok=True)


def warn_out_extra_files():
    ensure_dirs()
    allowed = {os.path.basename(OUTPUT_XLSX), os.path.basename(OUTPUT_JSON)}
    extras = []

    for filename in os.listdir(OUT_DIR):
        path = os.path.join(OUT_DIR, filename)

        if not os.path.isfile(path):
            continue

        if filename in allowed or filename in IGNORED_REPO_FILES:
            continue

        extras.append(filename)

    if extras:
        print("ATTENZIONE: in out/ ci sono file extra non creati da questi programmi:")
        for f in extras:
            print(f"  - {f}")
        print("Da ora questi programmi scrivono in out/ solo Excel e JSON finale.")



# ============================================================
# LOG PER RUN
# ============================================================
# Ogni esecuzione scrive sia a console sia in un file:
#   logs/run_YYYYMMDD_HHMMSS_<script>.log
# Il timestamp del log viene condiviso con la snapshot di backup della stessa run.

ORIGINAL_STDOUT = sys.stdout
ORIGINAL_STDERR = sys.stderr
CURRENT_LOG_PATH = None
CURRENT_LOG_FILE = None
LOGGING_STARTED = False


class TeeStream:
    def __init__(self, primary, secondary):
        self.primary = primary
        self.secondary = secondary
        self.encoding = getattr(primary, "encoding", "utf-8")

    def write(self, data):
        self.primary.write(data)
        self.primary.flush()
        self.secondary.write(data)
        self.secondary.flush()

    def flush(self):
        self.primary.flush()
        self.secondary.flush()

    def isatty(self):
        return False


def restore_logging():
    global CURRENT_LOG_FILE, LOGGING_STARTED

    if LOGGING_STARTED:
        sys.stdout = ORIGINAL_STDOUT
        sys.stderr = ORIGINAL_STDERR
        LOGGING_STARTED = False

    if CURRENT_LOG_FILE is not None:
        try:
            CURRENT_LOG_FILE.flush()
            CURRENT_LOG_FILE.close()
        except Exception:
            pass
        CURRENT_LOG_FILE = None


def start_run_logging(script_name):
    """
    Avvia il log della run.

    Da chiamare all'inizio di build/update/sync. Dopo questa chiamata, tutti i
    print() e gli errori non catturati finiscono anche nel file di log.
    """
    global CURRENT_LOG_PATH, CURRENT_LOG_FILE, LOGGING_STARTED

    ensure_dirs()

    if LOGGING_STARTED:
        return CURRENT_LOG_PATH

    safe_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", str(script_name)).strip("_") or "run"
    ts = get_run_backup_timestamp()
    CURRENT_LOG_PATH = os.path.join(LOG_DIR, f"run_{ts}_{safe_name}.log")

    CURRENT_LOG_FILE = open(CURRENT_LOG_PATH, "a", encoding="utf-8", buffering=1)
    sys.stdout = TeeStream(ORIGINAL_STDOUT, CURRENT_LOG_FILE)
    sys.stderr = TeeStream(ORIGINAL_STDERR, CURRENT_LOG_FILE)
    LOGGING_STARTED = True
    atexit.register(restore_logging)

    print("=" * 80)
    print(f"Log run: {CURRENT_LOG_PATH}")
    print(f"Script: {script_name}")
    print(f"Project root: {PROJECT_ROOT}")
    print(f"Started at: {datetime.now().isoformat(timespec='seconds')}")
    print("=" * 80)

    return CURRENT_LOG_PATH

# ============================================================
# BACKUP PER RUN TIMESTAMPATA
# ============================================================
# Nuova regola:
#   bkp/
#     YYYYMMDD_HHMMSS/
#       out/
#         one_piece_collection.xlsx
#         one_piece_collection.json
#       json/
#         products_singles_18.json
#         products_nonsingles_18.json
#         price_guide_18.json
#       manifest.json
#
# I file copiati/spostati mantengono il nome originale.
# Il timestamp sta nella cartella della run, non nel nome file.

RUN_BACKUP_TIMESTAMP = None
RUN_BACKUP_ENTRIES = []


def backup_timestamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def get_run_backup_timestamp(timestamp=None):
    global RUN_BACKUP_TIMESTAMP

    if timestamp:
        RUN_BACKUP_TIMESTAMP = timestamp
        return RUN_BACKUP_TIMESTAMP

    if RUN_BACKUP_TIMESTAMP is None:
        RUN_BACKUP_TIMESTAMP = backup_timestamp()

    return RUN_BACKUP_TIMESTAMP


def get_run_backup_dir(timestamp=None):
    ts = get_run_backup_timestamp(timestamp)
    return os.path.join(BKP_DIR, ts)


def path_is_inside(path, folder):
    try:
        return os.path.commonpath([os.path.abspath(path), os.path.abspath(folder)]) == os.path.abspath(folder)
    except Exception:
        return False


def backup_subdir_for_path(path):
    """
    Spostiamo/copiamo in backup solo file dalle cartelle finali:
    - out/  -> bkp/<timestamp>/out/
    - json/ -> bkp/<timestamp>/json/
    Gli intermedi in stg/ non vengono archiviati automaticamente.
    """
    if path_is_inside(path, OUT_DIR):
        return "out"
    if path_is_inside(path, JSON_DIR):
        return "json"
    return None


def write_backup_manifest(timestamp=None):
    ts = get_run_backup_timestamp(timestamp)
    backup_dir = get_run_backup_dir(ts)
    os.makedirs(backup_dir, exist_ok=True)

    manifest = {
        "timestamp": ts,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "projectRoot": PROJECT_ROOT,
        "backupRoot": backup_dir,
        "logFile": CURRENT_LOG_PATH,
        "policy": {
            "layout": "bkp/<timestamp>/{out,json}/<original filename>",
            "versioning": "No numeric versions; the timestamp is the folder name.",
            "scope": "Only files from out/ and json/ are backed up automatically. stg/ is regenerated technical output.",
        },
        "files": RUN_BACKUP_ENTRIES,
    }

    manifest_path = os.path.join(backup_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    return manifest_path


def backup_file(path, move=False, timestamp=None):
    """
    Copia o sposta un file in una snapshot di backup della run.

    Esempi:
      out/one_piece_collection.xlsx -> bkp/20260701_204512/out/one_piece_collection.xlsx
      json/price_guide_18.json      -> bkp/20260701_204512/json/price_guide_18.json
    """
    if not os.path.exists(path):
        return None

    subdir = backup_subdir_for_path(path)
    if subdir is None:
        print(f"Backup saltato: {path} non è in out/ o json/.")
        return None

    ts = get_run_backup_timestamp(timestamp)
    destination_dir = os.path.join(get_run_backup_dir(ts), subdir)
    os.makedirs(destination_dir, exist_ok=True)

    destination = os.path.join(destination_dir, os.path.basename(path))

    # Se lo stesso file è già stato archiviato nella stessa run, non lo dupli co
    # e non cambio nome: la snapshot deve restare leggibile.
    if os.path.exists(destination):
        print(f"Backup già presente nella run corrente, salto duplicato: {destination}")
        return destination

    size_bytes = os.path.getsize(path) if os.path.exists(path) else None

    if move:
        shutil.move(path, destination)
        action = "spostato"
    else:
        shutil.copy2(path, destination)
        action = "copiato"

    entry = {
        "source": path,
        "destination": destination,
        "folder": subdir,
        "action": action,
        "sizeBytes": size_bytes,
        "backedUpAt": datetime.now().isoformat(timespec="seconds"),
    }
    RUN_BACKUP_ENTRIES.append(entry)
    manifest_path = write_backup_manifest(ts)

    print(f"Backup {action}: {path} -> {destination}")
    print(f"Manifest backup: {manifest_path}")

    return destination


def backup_out_files(move=False, timestamp=None):
    ensure_dirs()
    ts = get_run_backup_timestamp(timestamp)
    backed_up = []

    if not os.path.isdir(OUT_DIR):
        return backed_up

    for filename in os.listdir(OUT_DIR):
        if filename in IGNORED_REPO_FILES:
            continue

        path = os.path.join(OUT_DIR, filename)
        if os.path.isfile(path):
            try:
                dest = backup_file(path, move=move, timestamp=ts)
                if dest:
                    backed_up.append(dest)
            except PermissionError:
                print(f"ATTENZIONE: file aperto/bloccato, non posso fare backup: {path}")

    return backed_up


def backup_json_files(move=True, timestamp=None):
    ensure_dirs()
    ts = get_run_backup_timestamp(timestamp)
    backed_up = []

    if not os.path.isdir(JSON_DIR):
        return backed_up

    for filename in os.listdir(JSON_DIR):
        path = os.path.join(JSON_DIR, filename)
        if os.path.isfile(path) and filename.lower().endswith(".json"):
            try:
                dest = backup_file(path, move=move, timestamp=ts)
                if dest:
                    backed_up.append(dest)
            except PermissionError:
                print(f"ATTENZIONE: file aperto/bloccato, non posso fare backup: {path}")

    return backed_up


def backup_known_files(move=False):
    """
    Backup iniziale di run.
    Archivia solo file finali out/ e JSON correnti json/ nella stessa snapshot:
      bkp/<timestamp>/out/
      bkp/<timestamp>/json/
    """
    ts = get_run_backup_timestamp()
    print("Backup file finali e JSON correnti...")
    print(f"Snapshot backup: {os.path.join(BKP_DIR, ts)}")
    backup_out_files(move=move, timestamp=ts)
    backup_json_files(move=move, timestamp=ts)
    write_backup_manifest(ts)

def download_file(url, destination, timeout=120):
    ensure_dirs()
    temp_destination = destination + ".download"

    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36",
            "Accept": "application/json,*/*",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            status = getattr(response, "status", 200)
            if status >= 400:
                raise RuntimeError(f"HTTP {status}")
            data = response.read()
    except urllib.error.URLError as e:
        raise RuntimeError(f"Download fallito da {url}: {e}") from e

    if not data:
        raise RuntimeError(f"Download vuoto da {url}")

    try:
        parsed = json.loads(data.decode("utf-8"))
    except Exception as e:
        raise RuntimeError(f"Il file scaricato non è JSON valido: {url}") from e

    with open(temp_destination, "w", encoding="utf-8") as f:
        json.dump(parsed, f, ensure_ascii=False)

    os.replace(temp_destination, destination)

    return destination


def fetch_text(url, timeout=60):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )

    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            status = getattr(response, "status", 200)
            if status >= 400:
                raise RuntimeError(f"HTTP {status}")
            raw = response.read()
    except urllib.error.URLError as e:
        raise RuntimeError(f"Non riesco a leggere la pagina {url}: {e}") from e

    return raw.decode("utf-8", errors="replace")


def extract_links_from_html(html, base_url):
    """
    Estrae link HTML in modo leggero, senza dipendere da BeautifulSoup.
    Ritorna una lista di dict: {text, href}.
    """
    links = []

    # Pattern volutamente pratico per le pagine Cardmarket Data Tables.
    pattern = re.compile(
        r'<a\b[^>]*?href=["\\\'](?P<href>[^"\\\']+)["\\\'][^>]*>(?P<text>.*?)</a>',
        flags=re.IGNORECASE | re.DOTALL,
    )

    for match in pattern.finditer(html):
        href = unescape(match.group("href")).strip()
        text_html = match.group("text")
        text = re.sub(r"<[^>]+>", " ", text_html)
        text = unescape(re.sub(r"\s+", " ", text)).strip()
        if href:
            links.append({"text": text, "href": urljoin(base_url, href)})

    return links


def fallback_cardmarket_one_piece_json_urls(reason):
    """
    Cardmarket a volte risponde 403 sulle pagine HTML Data Tables.
    In quel caso usiamo gli URL S3 pubblici noti e validiamo i JSON scaricati.
    """
    print("ATTENZIONE: non riesco a leggere le pagine HTML Cardmarket per scoprire i link JSON.")
    print(f"Motivo: {reason}")
    print("Uso fallback sugli URL S3 pubblici One Piece e poi valido i file scaricati.")

    labels = {
        "singles": "One Piece Singles fallback",
        "nonsingles": "One Piece Non-Singles fallback",
        "priceguide": "One Piece price guide fallback",
    }

    discovered = {}

    for key, url in CARDMARKET_FALLBACK_JSON_URLS.items():
        discovered[key] = {
            "text": labels[key],
            "href": url,
            "source": "fallback_s3",
            "reason": str(reason),
        }

    print("Link JSON One Piece usati in fallback:")
    for key, item in discovered.items():
        print(f"  {key}: {item['text']} -> {item['href']}")

    return discovered


def get_cardmarket_one_piece_json_urls():
    """Usa direttamente gli URL S3 pubblici dei JSON Cardmarket One Piece.

    Non legge più le pagine HTML Cardmarket, quindi evita i 403 e rende il run
    più semplice e prevedibile. I file vengono comunque validati dopo il download.
    """
    print("Uso direttamente gli URL JSON Cardmarket One Piece configurati:")
    labels = {
        "singles": "One Piece Singles",
        "nonsingles": "One Piece Non-Singles",
        "priceguide": "One Piece Price Guide",
    }
    discovered = {}
    for key, url in CARDMARKET_FALLBACK_JSON_URLS.items():
        discovered[key] = {
            "text": labels[key],
            "href": url,
            "source": "direct_s3",
            "reason": "configured_direct_url",
        }
        print(f"  {key}: {url}")
    return discovered


def validate_cardmarket_json_files():
    singles_path = find_json(PRODUCTS_SINGLES_PATTERN)
    nonsingles_path = find_json(PRODUCTS_NONSINGLES_PATTERN)
    price_path = find_json(PRICE_GUIDE_PATTERN)

    with open(singles_path, "r", encoding="utf-8") as f:
        singles = json.load(f)
    with open(nonsingles_path, "r", encoding="utf-8") as f:
        nonsingles = json.load(f)
    with open(price_path, "r", encoding="utf-8") as f:
        price = json.load(f)

    singles_products = singles.get("products", [])
    nonsingles_products = nonsingles.get("products", [])
    price_guides = price.get("priceGuides", [])

    if not singles_products:
        raise RuntimeError(f"{singles_path} non contiene products.")
    if not nonsingles_products:
        raise RuntimeError(f"{nonsingles_path} non contiene products.")
    if not price_guides:
        raise RuntimeError(f"{price_path} non contiene priceGuides.")

    singles_one_piece = sum(
        1 for p in singles_products
        if "one piece" in str(p.get("categoryName", "")).lower()
    )
    nonsingles_one_piece = sum(
        1 for p in nonsingles_products
        if "one piece" in str(p.get("categoryName", "")).lower()
    )

    if singles_one_piece == 0:
        raise RuntimeError(f"{singles_path} non sembra essere il catalogo Singles di One Piece.")
    if nonsingles_one_piece == 0:
        raise RuntimeError(f"{nonsingles_path} non sembra essere il catalogo Non-Singles di One Piece.")

    singles_ids = {p.get("idProduct") for p in singles_products if p.get("idProduct") is not None}
    price_ids = {p.get("idProduct") for p in price_guides if p.get("idProduct") is not None}
    overlap = len(singles_ids.intersection(price_ids))

    if overlap == 0:
        raise RuntimeError(
            f"{price_path} non sembra compatibile con {singles_path}: nessun idProduct in comune."
        )

    print(
        "JSON Cardmarket validati: "
        f"Singles One Piece={singles_one_piece}, "
        f"Non-Singles One Piece={nonsingles_one_piece}, "
        f"idProduct prezzi in comune={overlap}."
    )


def download_cardmarket_jsons(force=True):
    ensure_dirs()

    discovered = get_cardmarket_one_piece_json_urls()

    if force:
        # Prima di scaricare i nuovi JSON, sposto quelli correnti in bkp/json/.
        # Il nome resta originale + timestamp, senza versioni tipo v001.
        existing_jsons = [
            f for f in os.listdir(JSON_DIR)
            if os.path.isfile(os.path.join(JSON_DIR, f)) and f.lower().endswith(".json")
        ]
        if existing_jsons:
            print("Sposto i JSON correnti in backup prima del nuovo download...")
            backup_json_files(move=True)

    manifest = {
        "discoveredAt": datetime.now().isoformat(timespec="seconds"),
        "dataPages": CARDMARKET_DATA_PAGES,
        "files": {},
    }

    print("Scarico JSON Cardmarket One Piece in json/...")

    for key, item in discovered.items():
        url = item["href"]
        filename = os.path.basename(url.split("?")[0])
        destination = os.path.join(JSON_DIR, filename)

        if os.path.exists(destination) and not force:
            print(f"  Già presente: {destination}")
        else:
            print(f"  Uso {key}: {item['text']}")
            print(f"    URL:  {url}")
            print(f"    File: {destination}")
            download_file(url, destination)
            print(f"    Salvato.")

        manifest["files"][key] = {
            "label": item["text"],
            "url": url,
            "filename": filename,
            "localPath": destination,
        }

    with open(CARDMARKET_JSON_SOURCES_JSON, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(f"Manifest sorgenti JSON: {CARDMARKET_JSON_SOURCES_JSON}")

    validate_cardmarket_json_files()

def find_json(pattern):
    matches = glob.glob(os.path.join(JSON_DIR, pattern))
    if not matches:
        raise FileNotFoundError(f"Nessun file trovato in {JSON_DIR}/ con pattern {pattern}")
    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return matches[0]


def normalize_code(code):
    if not isinstance(code, str):
        return ""
    code = code.strip().upper().replace(" ", "")
    m = re.match(r"^(P)-?(\d{3})$", code)
    if m:
        return f"P-{m.group(2)}"
    m = re.match(r"^(OP|ST|EB|PRB)-?(\d{2})-(\d{3})$", code)
    if m:
        return f"{m.group(1)}{m.group(2)}-{m.group(3)}"
    return code


def format_number_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text == "" or text.lower() == "nan":
        return ""
    if re.fullmatch(r"\d+", text):
        return text.zfill(3)
    return text


def split_card_id(card_id):
    card_id = normalize_code(card_id)
    if not card_id:
        return "", ""
    if card_id.startswith("DON-"):
        return "DON", ""
    parts = card_id.split("-")
    if len(parts) == 2:
        return parts[0], format_number_text(parts[1])
    return card_id, ""


def normalize_language(value):
    if not isinstance(value, str):
        return "EN"
    return "EN" if value.upper().strip() == "EN" else "JP"


def clean_numeric(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return value


def estimate_code_from_product_name(name):
    if not isinstance(name, str):
        return ""
    n = name.lower()
    for key, code in KNOWN_EXPANSIONS:
        if key in n:
            return code
    return ""


def infer_language_from_product_name(name):
    if not isinstance(name, str):
        return "EN"
    n = name.lower()
    if "japanese" in n or "non-english" in n or "asia region legal" in n or "chinese" in n:
        return "JP"
    return "EN"


def clean_expansion_name(name):
    if not isinstance(name, str):
        return ""
    value = name
    cleanup_patterns = [
        r"\s+Booster Box Case.*$", r"\s+Booster Box.*$", r"\s+Sleeved Booster.*$",
        r"\s+Booster.*$", r"\s+Deck Pack.*$", r"\s+Pre-Release Pack.*$",
        r"\s+Dash Pack.*$", r"\s+\(Non-English\).*$", r"\s+\(Japanese\).*$",
        r"\s+\(Asia Region Legal\).*$", r"\s+\(English Version\).*$"
    ]
    for pattern in cleanup_patterns:
        value = re.sub(pattern, "", value, flags=re.I)
    return value.strip()


def is_promo_like_product_name(name):
    if not isinstance(name, str):
        return False
    n = name.lower()
    return any(k in n for k in PROMO_KEYWORDS)


def classify_cardmarket_product(row):
    card_id = str(row.get("ID Carta", "")).upper().strip()
    category = str(row.get("categoryName", "")).lower()
    cm_expansion_name = str(row.get("CM Expansion Name", "")).lower()
    cm_expansion_product = str(row.get("CM Expansion Product", "")).lower()
    cm_expansion_code = str(row.get("CM Expansion Code", "")).upper().strip()
    text_blob = " ".join([category, cm_expansion_name, cm_expansion_product])
    if card_id.startswith("P-"):
        return "Promo"
    if "promo products" in category:
        return "Promo"
    if is_promo_like_product_name(text_blob):
        return "Promo"
    card_expansion, _ = split_card_id(card_id)
    if card_expansion and cm_expansion_code and card_expansion != cm_expansion_code:
        return "Promo"
    return "Standard"


def build_variant_label(row):
    lang = row.get("CM Expansion Language", "EN")
    product_type = row.get("CM Product Type", "Standard")
    try:
        n = int(row.get("Cardmarket Variante N", 1))
    except Exception:
        n = 1
    if product_type == "Promo":
        return f"Promo [{lang}]"
    if n == 1:
        return f"Base [{lang}]"
    return f"Parallel / Alt / Reprint {n - 1} [{lang}]"


def load_expansion_map(output_csv=CARDMARKET_EXPANSIONS_CSV):
    try:
        nonsingles_file = find_json(PRODUCTS_NONSINGLES_PATTERN)
    except FileNotFoundError:
        print("products_nonsingles non trovato: uso mappa espansioni vuota.")
        return {}

    print(f"Leggo espansioni Cardmarket da: {nonsingles_file}")
    with open(nonsingles_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    df = pd.DataFrame(data.get("products", []))
    if df.empty:
        return {}
    for col in ["idExpansion", "name", "categoryName", "idProduct", "dateAdded"]:
        if col not in df.columns:
            df[col] = ""
    df["Lingua stimata"] = df["name"].apply(infer_language_from_product_name)
    df["Codice stimato"] = df["name"].apply(estimate_code_from_product_name)
    df["Nome espansione stimato"] = df["name"].apply(clean_expansion_name)
    df["Is Promo Product"] = df.apply(lambda r: "promo products" in str(r.get("categoryName", "")).lower() or is_promo_like_product_name(r.get("name", "")), axis=1)

    def priority(row):
        name = str(row.get("name", "")).lower()
        score = 100 if row.get("Codice stimato") else 0
        if "booster box case" in name:
            score += 1
        elif "booster box" in name:
            score += 5
        elif "booster" in name or "starter deck" in name or "ultimate deck" in name or "ultra deck" in name:
            score += 10
        if is_promo_like_product_name(name):
            score -= 50
        return score

    df["_priority"] = df.apply(priority, axis=1)
    df = df.sort_values(["idExpansion", "_priority"], ascending=[True, False])
    ref = df.groupby("idExpansion", as_index=False).agg(
        **{
            "Nome espansione stimato": ("Nome espansione stimato", "first"),
            "Codice stimato": ("Codice stimato", "first"),
            "Lingua stimata": ("Lingua stimata", "first"),
            "categoryName": ("categoryName", "first"),
            "name": ("name", "first"),
            "idProduct": ("idProduct", "first"),
            "dateAdded": ("dateAdded", "first"),
            "Is Promo Expansion": ("Is Promo Product", "max")
        }
    )
    ref = ref.rename(columns={
        "name": "Prodotto riferimento",
        "categoryName": "Categoria riferimento",
        "idProduct": "idProduct riferimento",
        "dateAdded": "Data prodotto riferimento"
    })
    ref.to_csv(output_csv, index=False, encoding="utf-8-sig")
    expansion_map = {}
    for _, row in ref.iterrows():
        expansion_map[row["idExpansion"]] = {
            "Cardmarket Expansion Name": row.get("Nome espansione stimato", ""),
            "Cardmarket Expansion Code": row.get("Codice stimato", ""),
            "Cardmarket Expansion Language": row.get("Lingua stimata", "EN"),
            "Cardmarket Expansion Product": row.get("Prodotto riferimento", ""),
            "Cardmarket Is Promo Expansion": bool(row.get("Is Promo Expansion", False)),
        }
    return expansion_map


def load_cardmarket_prices(output_csv=CARDMARKET_MERGED_CSV):
    if AUTO_DOWNLOAD_CARDMARKET_JSONS:
        download_cardmarket_jsons(force=True)
    expansion_map = load_expansion_map(CARDMARKET_EXPANSIONS_CSV)
    price_file = find_json(PRICE_GUIDE_PATTERN)
    products_file = find_json(PRODUCTS_SINGLES_PATTERN)
    print(f"Leggo price guide da: {price_file}")
    print(f"Leggo products singles da: {products_file}")
    with open(price_file, "r", encoding="utf-8") as f:
        price_data = json.load(f)
    with open(products_file, "r", encoding="utf-8") as f:
        products_data = json.load(f)
    price_created_at = price_data.get("createdAt", "")
    products_created_at = products_data.get("createdAt", "")
    df_price = pd.DataFrame(price_data.get("priceGuides", []))
    df_products = pd.DataFrame(products_data.get("products", []))
    if df_price.empty:
        raise ValueError("Il file price guide non contiene priceGuides.")
    if df_products.empty:
        raise ValueError("Il file products singles non contiene products.")

    def extract_code_from_name(name):
        if not isinstance(name, str):
            return ""
        matches = re.findall(r"\(([A-Z]{1,5}-?\d{2}-\d{3}|P-?\d{3})\)", name)
        return normalize_code(matches[-1]) if matches else ""

    def extract_clean_name(name):
        if not isinstance(name, str):
            return ""
        return re.sub(r"\s*\(([A-Z]{1,5}-?\d{2}-\d{3}|P-?\d{3})\)\s*$", "", name).strip()

    df_products["Cardmarket Nome"] = df_products["name"].apply(extract_clean_name)
    df_products["ID Carta"] = df_products["name"].apply(extract_code_from_name)
    df_products["Espansione"] = df_products["ID Carta"].apply(lambda x: split_card_id(x)[0])
    df_products["Numero"] = df_products["ID Carta"].apply(lambda x: split_card_id(x)[1])
    df_products["Numero"] = df_products["Numero"].apply(format_number_text)
    df = df_products.merge(df_price, on="idProduct", how="left", suffixes=("", "_price"))
    wanted = ["idProduct", "idCategory", "categoryName", "idExpansion", "idMetacard", "Cardmarket Nome", "ID Carta", "Espansione", "Numero", "avg", "low", "trend", "avg1", "avg7", "avg30", "dateAdded"]
    for c in wanted:
        if c not in df.columns:
            df[c] = ""
    df = df[wanted].copy()
    for c in ["avg", "low", "trend", "avg1", "avg7", "avg30"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["Cardmarket Price Created At"] = price_created_at
    df["Cardmarket Products Created At"] = products_created_at
    df = df[df["ID Carta"].notna() & (df["ID Carta"].astype(str).str.strip() != "")].copy()

    def get_info(id_expansion, key):
        return expansion_map.get(id_expansion, {}).get(key, "")

    df["CM Expansion Name"] = df["idExpansion"].apply(lambda x: get_info(x, "Cardmarket Expansion Name"))
    df["CM Expansion Code"] = df["idExpansion"].apply(lambda x: get_info(x, "Cardmarket Expansion Code"))
    df["CM Expansion Language"] = df["idExpansion"].apply(lambda x: get_info(x, "Cardmarket Expansion Language"))
    df["CM Expansion Product"] = df["idExpansion"].apply(lambda x: get_info(x, "Cardmarket Expansion Product"))
    df["CM Is Promo Expansion"] = df["idExpansion"].apply(lambda x: get_info(x, "Cardmarket Is Promo Expansion"))
    df["CM Expansion Language"] = df["CM Expansion Language"].replace("", "EN").fillna("EN").apply(normalize_language)
    df = df[df["CM Expansion Language"].isin(["EN", "JP"])].copy()
    df["CM Product Type"] = df.apply(classify_cardmarket_product, axis=1)

    diagnostic = df.copy()
    if EXCLUDE_PROMOS:
        before = len(df)
        df = df[df["CM Product Type"] != "Promo"].copy()
        print(f"Promo escluse dal catalogo prezzi: {before - len(df)}")

    price_col = PRICE_SOURCE_COLUMN if PRICE_SOURCE_COLUMN in df.columns else "trend"
    df["_sort_price"] = pd.to_numeric(df[price_col], errors="coerce").fillna(999999)
    df = df.sort_values(["ID Carta", "CM Expansion Language", "_sort_price", "idProduct"]).copy()
    df["Cardmarket Variante N"] = df.groupby(["ID Carta", "CM Expansion Language"]).cumcount() + 1
    df["Cardmarket Prodotti per carta"] = df.groupby(["ID Carta", "CM Expansion Language"])["idProduct"].transform("count")
    df["Variante"] = df.apply(build_variant_label, axis=1)
    df = df.drop(columns=["_sort_price"])
    diagnostic.to_csv(output_csv, index=False, encoding="utf-8-sig")
    print(f"Cardmarket diagnostico/intermedio: {output_csv}")
    return df


def save_final_json_from_df(df, source):
    data = df.copy().where(pd.notna(df), "")
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "priceSourceColumn": PRICE_SOURCE_COLUMN,
        "rows": len(data),
        "cards": data.to_dict(orient="records"),
    }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"JSON finale: {OUTPUT_JSON}")


def apply_qty_conditional_format(ws, qty_col, first_row=2):
    max_row = ws.max_row
    qty_missing_fill = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")
    qty_owned_fill = PatternFill(fill_type="solid", start_color="D9EAD3", end_color="D9EAD3")
    qty_letter = get_column_letter(qty_col)
    qty_range = f"{qty_letter}{first_row}:{qty_letter}{max_row}"
    ws.conditional_formatting.add(qty_range, FormulaRule(formula=[f'OR(${qty_letter}{first_row}=0,${qty_letter}{first_row}="")'], fill=qty_missing_fill))
    ws.conditional_formatting.add(qty_range, FormulaRule(formula=[f'${qty_letter}{first_row}>0'], fill=qty_owned_fill))



# ============================================================
# DATI UFFICIALI JP
# ============================================================
# Il sito JP/Asia può avere rarità diverse dal catalogo EN.
# Non usiamo override hardcoded: leggiamo il raw JP salvato da scrape_jp_official_raw().
# Se per lo stesso ID il sito JP contiene più stampe, scegliamo la rarità con priorità più alta
# fra quelle trovate ufficialmente. Esempio: OP16-042 può comparire come R e TR,
# quindi per le righe JP viene scelta TR perché presente nel catalogo JP.

JP_RARITY_PRIORITY = {
    "TR": 100,
    "SP": 95,
    "SEC": 90,
    "L": 80,
    "SR": 70,
    "R": 60,
    "UC": 50,
    "C": 40,
    "DON!!": 10,
}


def _rarity_priority(value):
    return JP_RARITY_PRIORITY.get(str(value).strip().upper(), 0)


def _load_jp_official_map():
    official_rows = {}

    if os.path.exists(JP_OFFICIAL_RAW_CSV):
        try:
            jp = pd.read_csv(JP_OFFICIAL_RAW_CSV, encoding="utf-8-sig", dtype={"ID Carta": str, "Numero": str}).fillna("")
            for _, row in jp.iterrows():
                card_id = normalize_code(str(row.get("ID Carta", "")).strip())
                rarity = str(row.get("Rarità", "")).strip()
                if not card_id or not rarity:
                    continue
                official_rows.setdefault(card_id, []).append({
                    "Rarità JP ufficiale": rarity,
                    "Nome JP ufficiale": str(row.get("Nome", "")).strip(),
                    "Tipo carta JP ufficiale": str(row.get("Tipo carta", "")).strip(),
                    "Color JP ufficiale": str(row.get("Color", "")).strip(),
                    "Fonte rarità JP": "Sito ufficiale JP",
                })
        except Exception as exc:
            print(f"ATTENZIONE: non riesco a leggere {JP_OFFICIAL_RAW_CSV}: {exc}")

    official = {}
    for card_id, rows in official_rows.items():
        rows = [r for r in rows if str(r.get("Rarità JP ufficiale", "")).strip()]
        if not rows:
            continue
        best = sorted(rows, key=lambda r: _rarity_priority(r.get("Rarità JP ufficiale", "")), reverse=True)[0].copy()
        candidates = sorted({str(r.get("Rarità JP ufficiale", "")).strip() for r in rows if str(r.get("Rarità JP ufficiale", "")).strip()}, key=_rarity_priority, reverse=True)
        best["Rarità JP candidate"] = ", ".join(candidates)
        if len(candidates) > 1:
            best["Fonte rarità JP"] = "Sito ufficiale JP, scelta priorità fra ristampe"
        official[card_id] = best

    return official


def apply_jp_official_corrections(df):
    data = df.copy()
    for col in ["Rarità JP ufficiale", "Rarità JP candidate", "Nome JP ufficiale", "Tipo carta JP ufficiale", "Color JP ufficiale", "Fonte rarità JP"]:
        if col not in data.columns:
            data[col] = ""

    official = _load_jp_official_map()
    if not official or "Lingua" not in data.columns or "ID Carta" not in data.columns:
        return data

    changed = 0
    for idx, row in data.iterrows():
        lang = normalize_language(str(row.get("Lingua", "EN")))
        if lang != "JP":
            continue
        card_id = normalize_code(str(row.get("ID Carta", "")).strip())
        info = official.get(card_id)
        if not info:
            continue
        rarity = str(info.get("Rarità JP ufficiale", "")).strip()
        if rarity:
            old = str(row.get("Rarità", "")).strip()
            data.at[idx, "Rarità"] = rarity
            data.at[idx, "Rarità JP ufficiale"] = rarity
            if old != rarity:
                changed += 1
        for src, dst in [
            ("Rarità JP candidate", "Rarità JP candidate"),
            ("Nome JP ufficiale", "Nome JP ufficiale"),
            ("Tipo carta JP ufficiale", "Tipo carta JP ufficiale"),
            ("Color JP ufficiale", "Color JP ufficiale"),
            ("Fonte rarità JP", "Fonte rarità JP"),
        ]:
            data.at[idx, dst] = info.get(src, "")
    if changed:
        print(f"Rarità JP corrette con dati ufficiali dal sito: {changed}")
    return data


def append_value_history(df, source):
    """Aggiunge un punto storico del valore collezione in stg/value_history.csv."""
    data = df.copy()
    for c in ["Quantità", "Valore", "Valore totale"]:
        if c not in data.columns:
            data[c] = 0
        data[c] = pd.to_numeric(data[c], errors="coerce").fillna(0)
    data["Valore totale"] = data["Quantità"] * data["Valore"]
    row = {
        "Data": datetime.now().isoformat(timespec="seconds"),
        "Fonte": source,
        "Carte": int(len(data)),
        "Carte possedute": int((data["Quantità"] > 0).sum()),
        "Quantità totale": int(data["Quantità"].sum()),
        "Valore collezione (€)": float(data["Valore totale"].sum()),
        "Valore database (€)": float(data["Valore"].sum()),
        "Carte con prezzo": int((data["Valore"] > 0).sum()),
    }
    ensure_dirs()
    if os.path.exists(VALUE_HISTORY_CSV):
        history = pd.read_csv(VALUE_HISTORY_CSV, encoding="utf-8-sig")
        history = pd.concat([history, pd.DataFrame([row])], ignore_index=True)
    else:
        history = pd.DataFrame([row])
    history.to_csv(VALUE_HISTORY_CSV, index=False, encoding="utf-8-sig")
    print(f"Storico valore aggiornato: {VALUE_HISTORY_CSV}")

# ============================================================
# DASHBOARD + CONFRONTO PREZZI
# ============================================================
# Questa sezione è volutamente centralizzata in common:
# build, update_values e sync devono generare SEMPRE lo stesso Excel.

PRICE_TREND_REPORT_CSV = os.path.join(STG_DIR, "price_trend_report.csv")
TOP_5_AUMENTI_CSV = os.path.join(STG_DIR, "top_5_aumenti.csv")
TOP_5_CALI_CSV = os.path.join(STG_DIR, "top_5_cali.csv")


def normalize_text_for_key(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", " ", text)


def normalize_variant_for_price_key(value):
    text = normalize_text_for_key(value).lower()
    text = text.replace(" - jp manuale", "")
    text = text.replace("jp manuale", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_idproduct_for_price_key(value):
    text = normalize_text_for_key(value)
    if text == "":
        return ""
    try:
        return str(int(float(text)))
    except Exception:
        return text


def find_latest_backup_output_json():
    """Trova l'ultimo bkp/<timestamp>/out/one_piece_collection.json."""
    ensure_dirs()
    candidates = []
    if not os.path.isdir(BKP_DIR):
        return None
    for name in os.listdir(BKP_DIR):
        snap = os.path.join(BKP_DIR, name)
        if not os.path.isdir(snap):
            continue
        candidate = os.path.join(snap, "out", os.path.basename(OUTPUT_JSON))
        if os.path.exists(candidate):
            candidates.append((name, os.path.getmtime(candidate), candidate))
    if not candidates:
        return None
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def load_collection_json_cards(path):
    if not path or not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        payload = json.load(f)
    if isinstance(payload, dict) and isinstance(payload.get("cards"), list):
        return payload["cards"]
    if isinstance(payload, list):
        return payload
    return []


def build_previous_price_maps(cards):
    by_exact = {}
    by_variant = {}
    by_card_lang = {}

    for row in cards:
        card_id = normalize_code(normalize_text_for_key(row.get("ID Carta", "")))
        lang = normalize_language(normalize_text_for_key(row.get("Lingua", "EN")))
        variant = normalize_variant_for_price_key(row.get("Variante", ""))
        idp = normalize_idproduct_for_price_key(row.get("Cardmarket idProduct", ""))
        value = pd.to_numeric(row.get("Valore", 0), errors="coerce")
        if pd.isna(value):
            value = 0.0
        value = float(value)

        if not card_id:
            continue

        if idp:
            by_exact[(card_id, lang, idp)] = value

        by_variant.setdefault((card_id, lang, variant), []).append(value)
        by_card_lang.setdefault((card_id, lang), []).append(value)

    unique_variant = {k: v[0] for k, v in by_variant.items() if len(v) == 1}
    unique_card_lang = {k: v[0] for k, v in by_card_lang.items() if len(v) == 1}
    return by_exact, unique_variant, unique_card_lang


def add_price_trends_from_latest_backup(df, previous_json_path=None):
    """
    Aggiunge confronto prezzo rispetto all'ultimo JSON finale salvato in backup.
    Colonne aggiunte:
    - Valore precedente
    - Variazione valore
    - Variazione %
    - Trend prezzo
    - Data confronto
    - Fonte confronto
    """
    out = df.copy()

    for col in ["Valore", "Cardmarket idProduct", "ID Carta", "Lingua", "Variante", "Quantità"]:
        if col not in out.columns:
            out[col] = ""

    previous_json_path = previous_json_path or find_latest_backup_output_json()
    comparison_time = datetime.now().isoformat(timespec="seconds")

    # Il confronto prezzi è sensato solo per le carte possedute.
    # Le carte con Quantità = 0 restano fuori dai trend, top aumenti/cali e delta netto.
    if "Quantità" not in out.columns:
        out["Quantità"] = 0
    owned_mask = pd.to_numeric(out["Quantità"], errors="coerce").fillna(0) > 0

    if not previous_json_path:
        out["Valore precedente"] = ""
        out["Variazione valore"] = ""
        out["Variazione %"] = ""
        out["Trend prezzo"] = "Non posseduta"
        out.loc[owned_mask, "Trend prezzo"] = "Nessun confronto"
        out["Data confronto"] = comparison_time
        out["Fonte confronto"] = "Confronto eseguito solo sulle carte possedute. Nessun JSON finale precedente in bkp/."
        print("Confronto prezzi: nessun JSON finale precedente trovato in bkp/.")
        print(f"Carte possedute confrontabili: {int(owned_mask.sum())} / {len(out)}")
        return out

    cards = load_collection_json_cards(previous_json_path)
    by_exact, by_variant, by_card_lang = build_previous_price_maps(cards)

    prev_values = []
    deltas = []
    pct_deltas = []
    trends = []
    sources = []

    for _, row in out.iterrows():
        card_id = normalize_code(normalize_text_for_key(row.get("ID Carta", "")))
        lang = normalize_language(normalize_text_for_key(row.get("Lingua", "EN")))
        variant = normalize_variant_for_price_key(row.get("Variante", ""))
        idp = normalize_idproduct_for_price_key(row.get("Cardmarket idProduct", ""))
        current = pd.to_numeric(row.get("Valore", 0), errors="coerce")
        if pd.isna(current):
            current = 0.0
        current = float(current)
        qty = pd.to_numeric(row.get("Quantità", 0), errors="coerce")
        if pd.isna(qty):
            qty = 0
        if float(qty) <= 0:
            prev_values.append("")
            deltas.append("")
            pct_deltas.append("")
            trends.append("Non posseduta")
            sources.append("Confronto escluso: carta non posseduta")
            continue

        previous = None
        source = ""

        if idp and (card_id, lang, idp) in by_exact:
            previous = by_exact[(card_id, lang, idp)]
            source = "Match idProduct"
        elif (card_id, lang, variant) in by_variant:
            previous = by_variant[(card_id, lang, variant)]
            source = "Match carta+lingua+variante"
        elif (card_id, lang) in by_card_lang:
            previous = by_card_lang[(card_id, lang)]
            source = "Match unico carta+lingua"

        if previous is None:
            prev_values.append("")
            deltas.append("")
            pct_deltas.append("")
            trends.append("Nuova / non confrontata")
            sources.append(previous_json_path)
            continue

        delta = current - previous
        if previous != 0:
            pct = delta / previous
        else:
            pct = "" if current == 0 else ""

        if delta > 0:
            trend = "In aumento"
        elif delta < 0:
            trend = "In calo"
        else:
            trend = "Stabile"

        prev_values.append(float(previous))
        deltas.append(float(delta))
        pct_deltas.append(pct)
        trends.append(trend)
        sources.append(f"{source}: {previous_json_path}")

    out["Valore precedente"] = prev_values
    out["Variazione valore"] = deltas
    out["Variazione %"] = pct_deltas
    out["Trend prezzo"] = trends
    out["Data confronto"] = comparison_time
    out["Fonte confronto"] = sources

    compared = sum(1 for x in trends if x in ["In aumento", "In calo", "Stabile"])
    owned = sum(1 for x in trends if x != "Non posseduta")
    print(f"Confronto prezzi: usato JSON precedente {previous_json_path}")
    print(f"Carte possedute confrontate: {compared} / {owned}. Carte non possedute escluse: {len(out) - owned}")
    return out


def save_price_trend_reports(df):
    data = df.copy()
    if "Variazione valore" not in data.columns:
        return
    if "Quantità" not in data.columns:
        data["Quantità"] = 0
    data = data[pd.to_numeric(data["Quantità"], errors="coerce").fillna(0) > 0].copy()
    data["_delta"] = pd.to_numeric(data["Variazione valore"], errors="coerce")
    data["_abs_delta"] = data["_delta"].abs()
    data.to_csv(PRICE_TREND_REPORT_CSV, index=False, encoding="utf-8-sig")

    cols = [c for c in ["ID Carta", "Espansione", "Numero", "Nome", "Lingua", "Variante", "Quantità", "Valore precedente", "Valore", "Variazione valore", "Variazione %", "Trend prezzo"] if c in data.columns]
    top_up = data[data["_delta"] > 0].sort_values("_delta", ascending=False).head(5)[cols]
    top_down = data[data["_delta"] < 0].sort_values("_delta", ascending=True).head(5)[cols]
    top_up.to_csv(TOP_5_AUMENTI_CSV, index=False, encoding="utf-8-sig")
    top_down.to_csv(TOP_5_CALI_CSV, index=False, encoding="utf-8-sig")
    print(f"Report trend prezzi: {PRICE_TREND_REPORT_CSV}")
    print(f"Top 5 aumenti: {TOP_5_AUMENTI_CSV}")
    print(f"Top 5 cali: {TOP_5_CALI_CSV}")


def _excel_safe_df(df):
    data = df.copy()
    for c in data.columns:
        if data[c].dtype == "object":
            data[c] = data[c].apply(clean_for_excel)
    return data


def _currency_format(ws, col, start_row=2):
    if not col:
        return
    for r in range(start_row, ws.max_row + 1):
        ws.cell(row=r, column=col).number_format = '€ #,##0.00'


def _percent_format(ws, col, start_row=2):
    if not col:
        return
    for r in range(start_row, ws.max_row + 1):
        ws.cell(row=r, column=col).number_format = '0.00%'


def _write_summary_table(dash, col_map, start_row, start_col, title, items, field_name, max_row):
    dash.cell(row=start_row, column=start_col).value = title
    dash.cell(row=start_row, column=start_col).font = Font(size=13, bold=True, color="1F4E78")
    headers = [field_name, "Carte", "Possedute", "Quantità", "Valore"]
    for i, h in enumerate(headers):
        cell = dash.cell(row=start_row + 1, column=start_col + i)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center")
    field_letter = get_column_letter(col_map[field_name])
    qty_letter = get_column_letter(col_map["Quantità"])
    total_letter = get_column_letter(col_map["Valore totale"])
    field_range = f"'Carte'!${field_letter}$2:${field_letter}${max_row}"
    qty_range = f"'Carte'!${qty_letter}$2:${qty_letter}${max_row}"
    total_range = f"'Carte'!${total_letter}$2:${total_letter}${max_row}"
    for r, item in enumerate(items, start=start_row + 2):
        dash.cell(row=r, column=start_col).value = item
        item_ref = dash.cell(row=r, column=start_col).coordinate
        dash.cell(row=r, column=start_col + 1).value = f'=COUNTIF({field_range},{item_ref})'
        dash.cell(row=r, column=start_col + 2).value = f'=COUNTIFS({field_range},{item_ref},{qty_range},">0")'
        dash.cell(row=r, column=start_col + 3).value = f'=SUMIF({field_range},{item_ref},{qty_range})'
        dash.cell(row=r, column=start_col + 4).value = f'=SUMIF({field_range},{item_ref},{total_range})'
        dash.cell(row=r, column=start_col + 4).number_format = '€ #,##0.00'
    return start_row + 1, start_row + 1 + len(items)


def create_collection_workbook_with_dashboard(df, output_xlsx=OUTPUT_XLSX):
    """Crea l'Excel finale completo con Dashboard, KPI, grafici e trend prezzi."""
    data = df.copy()
    preferred = [
        "ID Carta", "Espansione", "Numero", "Nome", "Lingua", "Variante", "CM Product Type",
        "Rarità", "Rarità JP ufficiale", "Fonte rarità JP", "Tipo carta", "Nome JP ufficiale", "Tipo carta JP ufficiale", "Color", "Color JP ufficiale", "Quantità", "Valore", "Valore totale",
        "Valore precedente", "Variazione valore", "Variazione %", "Trend prezzo", "Data confronto", "Fonte confronto",
        "Fonte prezzo", "CM_Data_Prezzo", "Cardmarket idProduct", "Cardmarket Nome", "Cardmarket Prodotti per carta",
        "CM_Low", "CM_Trend", "CM_Avg", "CM_Avg1", "CM_Avg7", "CM_Avg30",
        "CM Expansion Name", "CM Expansion Code", "CM Expansion Language", "CM Expansion Product",
        "Cost", "Life", "Attribute", "Power", "Counter", "Block", "Type", "Effect", "Trigger", "Card Set(s)", "Notes",
        "Series ID", "Set Code Ricerca", "Set Label Ricerca"
    ]
    for c in preferred:
        if c not in data.columns:
            data[c] = ""
    data = _excel_safe_df(data[preferred].copy())
    data["Numero"] = data["Numero"].apply(format_number_text)
    data["Quantità"] = pd.to_numeric(data["Quantità"], errors="coerce").fillna(0).astype(int)
    data["Valore"] = pd.to_numeric(data["Valore"], errors="coerce").fillna(0.0)
    for c in ["Valore precedente", "Variazione valore", "Variazione %"]:
        data[c] = pd.to_numeric(data[c], errors="coerce")
    data["Valore totale"] = ""

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        data.to_excel(writer, index=False, sheet_name="Carte")

    wb = load_workbook(output_xlsx)
    ws = wb["Carte"]
    ws.freeze_panes = "A2"
    max_row, max_col = ws.max_row, ws.max_column
    col_map = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    formula_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    qty_col = col_map["Quantità"]
    value_col = col_map["Valore"]
    total_col = col_map["Valore totale"]
    numero_col = col_map["Numero"]

    for r in range(2, max_row + 1):
        ws.cell(r, numero_col).number_format = "@"
        ws.cell(r, value_col).fill = editable_fill
        ws.cell(r, total_col).fill = formula_fill
        ql = get_column_letter(qty_col)
        vl = get_column_letter(value_col)
        ws.cell(r, total_col).value = f"={ql}{r}*{vl}{r}"
        ws.cell(r, qty_col).number_format = "0"
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    for name in ["Valore", "Valore totale", "Valore precedente", "Variazione valore", "CM_Low", "CM_Trend", "CM_Avg", "CM_Avg1", "CM_Avg7", "CM_Avg30"]:
        _currency_format(ws, col_map.get(name))
    _percent_format(ws, col_map.get("Variazione %"))
    apply_qty_conditional_format(ws, qty_col)

    widths = {"A":16,"B":12,"C":10,"D":24,"E":8,"F":40,"G":16,"H":10,"I":14,"J":14,"K":10,"L":12,"M":14,"N":14,"O":14,"P":12,"Q":16,"R":20,"S":45,"T":38,"U":24,"V":16,"W":28,"X":16,"Y":12,"Z":12,"AA":12,"AB":12,"AC":12,"AD":12,"AE":24,"AF":14,"AG":16,"AH":34,"AI":8,"AJ":8,"AK":14,"AL":10,"AM":10,"AN":8,"AO":28,"AP":52,"AQ":34,"AR":34,"AS":30,"AT":12,"AU":16,"AV":38}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    dash = wb.create_sheet("Dashboard", 0)
    dash.sheet_view.showGridLines = False
    for col in range(1, 16):
        dash.column_dimensions[get_column_letter(col)].width = 18

    dash["A1"] = "Dashboard Collezione One Piece Card Game"
    dash["A1"].font = Font(size=18, bold=True, color="1F4E78")
    dash["A2"] = "KPI, grafici e confronto valore rispetto all'ultimo JSON finale salvato in bkp/."
    dash["A2"].font = Font(italic=True, color="666666")

    id_l = get_column_letter(col_map["ID Carta"])
    q_l = get_column_letter(col_map["Quantità"])
    t_l = get_column_letter(col_map["Valore totale"])
    v_l = get_column_letter(col_map["Valore"])
    lang_l = get_column_letter(col_map["Lingua"])
    trend_l = get_column_letter(col_map["Trend prezzo"])
    delta_l = get_column_letter(col_map["Variazione valore"])
    id_r = f"'Carte'!${id_l}$2:${id_l}${max_row}"
    q_r = f"'Carte'!${q_l}$2:${q_l}${max_row}"
    t_r = f"'Carte'!${t_l}$2:${t_l}${max_row}"
    v_r = f"'Carte'!${v_l}$2:${v_l}${max_row}"
    lang_r = f"'Carte'!${lang_l}$2:${lang_l}${max_row}"
    trend_r = f"'Carte'!${trend_l}$2:${trend_l}${max_row}"
    delta_r = f"'Carte'!${delta_l}$2:${delta_l}${max_row}"

    kpis = [
        ("A4", "Carte database", f'=COUNTA({id_r})'),
        ("C4", "Carte possedute", f'=COUNTIF({q_r},">0")'),
        ("E4", "Quantità totale", f'=SUM({q_r})'),
        ("G4", "Valore totale", f'=SUM({t_r})'),
        ("I4", "Carte EN", f'=COUNTIF({lang_r},"EN")'),
        ("K4", "Carte JP", f'=COUNTIF({lang_r},"JP")'),
        ("M4", "Carte con prezzo", f'=COUNTIF({v_r},">0")'),
        ("A7", "In aumento possedute", f'=COUNTIFS({trend_r},"In aumento",{q_r},">0")'),
        ("C7", "In calo possedute", f'=COUNTIFS({trend_r},"In calo",{q_r},">0")'),
        ("E7", "Stabili possedute", f'=COUNTIFS({trend_r},"Stabile",{q_r},">0")'),
        ("G7", "Nuove/non confrontate possedute", f'=COUNTIFS({trend_r},"Nuova / non confrontata",{q_r},">0")'),
        ("I7", "Somma aumenti possedute", f'=SUMIFS({delta_r},{delta_r},">0",{q_r},">0")'),
        ("K7", "Somma cali possedute", f'=SUMIFS({delta_r},{delta_r},"<0",{q_r},">0")'),
        ("M7", "Delta netto possedute", f'=SUMIFS({delta_r},{q_r},">0")'),
    ]
    for cell, label, formula in kpis:
        letters = re.match(r"([A-Z]+)", cell).group(1)
        val_cell = f"{letters}{int(re.search(r'\d+', cell).group()) + 1}"
        dash[cell] = label
        dash[cell].font = Font(bold=True, color="FFFFFF")
        dash[cell].fill = PatternFill("solid", fgColor="1F4E78")
        dash[cell].alignment = Alignment(horizontal="center")
        dash[val_cell] = formula
        dash[val_cell].font = Font(size=14, bold=True)
        dash[val_cell].fill = PatternFill("solid", fgColor="D9EAF7")
        dash[val_cell].alignment = Alignment(horizontal="center")
        if label in ["Valore totale", "Somma aumenti possedute", "Somma cali possedute", "Delta netto possedute"]:
            dash[val_cell].number_format = '€ #,##0.00'

    # Tabelle riassuntive
    rarities = sorted([x for x in data["Rarità"].dropna().unique().tolist() if str(x).strip()])
    colors = sorted([x for x in data["Color"].dropna().unique().tolist() if str(x).strip()])
    languages = sorted([x for x in data["Lingua"].dropna().unique().tolist() if str(x).strip()])
    expansions = sorted([x for x in data["Espansione"].dropna().unique().tolist() if str(x).strip()])
    trend_items = ["In aumento", "In calo", "Stabile", "Nuova / non confrontata", "Nessun confronto"]

    rarity_header, rarity_last = _write_summary_table(dash, col_map, 11, 1, "Riepilogo per rarità", rarities, "Rarità", max_row)
    color_header, color_last = _write_summary_table(dash, col_map, 11, 7, "Riepilogo per colore", colors, "Color", max_row)
    lang_header, lang_last = _write_summary_table(dash, col_map, max(rarity_last, color_last) + 4, 1, "Riepilogo per lingua", languages, "Lingua", max_row)
    exp_header, exp_last = _write_summary_table(dash, col_map, max(rarity_last, color_last) + 4, 7, "Riepilogo per espansione", expansions, "Espansione", max_row)

    trend_start = max(lang_last, exp_last) + 4
    dash.cell(row=trend_start, column=1).value = "Riepilogo andamento prezzi"
    dash.cell(row=trend_start, column=1).font = Font(size=13, bold=True, color="1F4E78")
    trend_headers = ["Trend prezzo", "Carte", "Delta totale"]
    for i, h in enumerate(trend_headers):
        cell = dash.cell(row=trend_start + 1, column=1 + i)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    for r, item in enumerate(trend_items, start=trend_start + 2):
        dash.cell(r, 1).value = item
        ref = dash.cell(r, 1).coordinate
        dash.cell(r, 2).value = f'=COUNTIFS({trend_r},{ref},{q_r},">0")'
        dash.cell(r, 3).value = f'=SUMIFS({delta_r},{trend_r},{ref},{q_r},">0")'
        dash.cell(r, 3).number_format = '€ #,##0.00'
    trend_header = trend_start + 1
    trend_last = trend_start + 1 + len(trend_items)

    top_start = trend_last + 4
    dash.cell(row=top_start, column=1).value = "Top 5 carte in aumento"
    dash.cell(row=top_start, column=1).font = Font(size=13, bold=True, color="1F4E78")
    dash.cell(row=top_start, column=7).value = "Top 5 carte in calo"
    dash.cell(row=top_start, column=7).font = Font(size=13, bold=True, color="1F4E78")

    top_cols = ["ID Carta", "Nome", "Lingua", "Valore precedente", "Valore", "Variazione valore"]
    for offset in [0, 6]:
        for i, h in enumerate(top_cols, start=1):
            cell = dash.cell(row=top_start + 1, column=offset + i)
            cell.value = h
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5B9BD5")

    top_df = data.copy()
    if "Quantità" in top_df.columns:
        top_df = top_df[pd.to_numeric(top_df["Quantità"], errors="coerce").fillna(0) > 0].copy()
    top_df["_delta"] = pd.to_numeric(top_df["Variazione valore"], errors="coerce")
    top_df = top_df[pd.to_numeric(top_df.get("Quantità", 0), errors="coerce").fillna(0) > 0].copy()
    top_up = top_df[top_df["_delta"] > 0].sort_values("_delta", ascending=False).head(5)
    top_down = top_df[top_df["_delta"] < 0].sort_values("_delta", ascending=True).head(5)
    for idx, (_, row) in enumerate(top_up.iterrows(), start=top_start + 2):
        for j, h in enumerate(top_cols, start=1):
            dash.cell(idx, j).value = row.get(h, "")
        for j in [4,5,6]:
            dash.cell(idx, j).number_format = '€ #,##0.00'
    for idx, (_, row) in enumerate(top_down.iterrows(), start=top_start + 2):
        for j, h in enumerate(top_cols, start=7):
            dash.cell(idx, j).value = row.get(h, "")
        for j in [10,11,12]:
            dash.cell(idx, j).number_format = '€ #,##0.00'

    valuable_start = top_start + 10
    dash.cell(row=valuable_start, column=1).value = "Top 20 carte più care nel database"
    dash.cell(row=valuable_start, column=1).font = Font(size=13, bold=True, color="1F4E78")
    val_headers = ["ID Carta", "Lingua", "Nome", "Variante", "Valore"]
    for i, h in enumerate(val_headers, start=1):
        cell = dash.cell(valuable_start + 1, i)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    high = data.copy()
    high["Valore"] = pd.to_numeric(high["Valore"], errors="coerce").fillna(0)
    for r, (_, row) in enumerate(high.sort_values("Valore", ascending=False).head(20).iterrows(), start=valuable_start + 2):
        dash.cell(r, 1).value = row.get("ID Carta", "")
        dash.cell(r, 2).value = row.get("Lingua", "")
        dash.cell(r, 3).value = row.get("Nome", "")
        dash.cell(r, 4).value = row.get("Variante", "")
        dash.cell(r, 5).value = row.get("Valore", 0)
        dash.cell(r, 5).number_format = '€ #,##0.00'

    # Grafici
    chart_anchor_row = valuable_start + 2
    if rarity_last > rarity_header:
        chart1 = BarChart()
        chart1.title = "Quantità possedute per rarità"
        chart1.y_axis.title = "Quantità"
        chart1.x_axis.title = "Rarità"
        chart1.height = 8
        chart1.width = 16
        chart1.add_data(Reference(dash, min_col=4, min_row=rarity_header, max_row=rarity_last), titles_from_data=True)
        chart1.set_categories(Reference(dash, min_col=1, min_row=rarity_header + 1, max_row=rarity_last))
        chart1.legend = None
        dash.add_chart(chart1, f"H{chart_anchor_row}")
    if trend_last > trend_header:
        chart2 = BarChart()
        chart2.title = "Andamento prezzi carte possedute"
        chart2.y_axis.title = "Carte"
        chart2.x_axis.title = "Trend"
        chart2.height = 8
        chart2.width = 16
        chart2.add_data(Reference(dash, min_col=2, min_row=trend_header, max_row=trend_last), titles_from_data=True)
        chart2.set_categories(Reference(dash, min_col=1, min_row=trend_header + 1, max_row=trend_last))
        chart2.legend = None
        dash.add_chart(chart2, f"H{chart_anchor_row + 16}")
    if lang_last > lang_header:
        chart3 = PieChart()
        chart3.title = "Valore per lingua"
        chart3.height = 8
        chart3.width = 12
        chart3.add_data(Reference(dash, min_col=5, min_row=lang_header, max_row=lang_last), titles_from_data=True)
        chart3.set_categories(Reference(dash, min_col=1, min_row=lang_header + 1, max_row=lang_last))
        dash.add_chart(chart3, f"H{chart_anchor_row + 32}")

    note_row = valuable_start + 25
    dash[f"A{note_row}"] = "Come usarlo"
    dash[f"A{note_row}"].font = Font(size=13, bold=True, color="1F4E78")
    notes = [
        "Modifica Quantità nel foglio Carte.",
        "La Dashboard viene rigenerata da build, update_values e sync.",
        "Il confronto prezzo usa l'ultimo out/one_piece_collection.json salvato in bkp/<timestamp>/out/.",
        "Il confronto valori viene calcolato solo sulle carte possedute, Quantità > 0.",
        "Aumenti, cali e delta prezzo vengono calcolati solo sulle carte possedute, cioè Quantità > 0.",
        "Se non esiste un JSON precedente, Trend prezzo mostra Nessun confronto.",
        "Gli intermedi e i report trend sono in stg/.",
        "I log sono in logs/."
    ]
    for i, note in enumerate(notes, start=note_row + 1):
        dash[f"A{i}"] = f"• {note}"

    for row in dash.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    dash.freeze_panes = "A11"
    wb.save(output_xlsx)
    print(f"Excel finale con Dashboard creato: {output_xlsx}")

