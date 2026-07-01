import os
import re
import time
import random
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from one_piece_common import *
from one_piece_collection_build import (
    BASE_URL,
    safe_get,
    get_lines_from_html,
    discover_sets,
    parse_result_count,
    parse_cards_from_lines,
    enrich,
    build_excel,
    sanitize_dataframe_for_excel,
)


# ============================================================
# PROGRAMMA 4
# Sync completo partendo dal file esistente.
#
# Cosa fa:
# - legge out/one_piece_collection.xlsx esistente
# - conserva le Quantità inserite a mano
# - interroga di nuovo il sito Bandai
# - legge di nuovo i JSON in json/
# - aggiorna prezzi/varianti/dati carta
# - aggiunge nuove carte/espansioni in fondo
# - salva solo Excel + JSON finale in out/
# - salva intermedi/report in stg/
# ============================================================

SYNC_RAW_CSV = os.path.join(STG_DIR, "bandai_cards_sync_raw.csv")
SYNC_FINAL_CSV = os.path.join(STG_DIR, "one_piece_collection_sync_stg.csv")
SYNC_REPORT_CSV = os.path.join(STG_DIR, "one_piece_sync_report.csv")

# Se True, fa backup di Excel e JSON finale prima di sovrascriverli.
BACKUP_FINAL_FILES = True

# Se True, le nuove righe finiscono in fondo, dopo le righe già presenti nel tuo Excel.
KEEP_EXISTING_ORDER_AND_APPEND_NEW = True

# Se True, per righe non matchate perfettamente prova a recuperare la Quantità con chiavi meno strette.
USE_FALLBACK_QUANTITY_MATCH = True


def get_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None and str(value).strip() != "":
            headers[str(value).strip()] = col
    return headers


def read_existing_excel():
    if not os.path.exists(OUTPUT_XLSX):
        raise FileNotFoundError(
            f"File non trovato: {OUTPUT_XLSX}. "
            "Prima crea il file con one_piece_collection_build.py."
        )

    wb = load_workbook(OUTPUT_XLSX, data_only=False)

    if "Carte" not in wb.sheetnames:
        raise ValueError("Nel file Excel non trovo il foglio 'Carte'.")

    ws = wb["Carte"]
    headers = get_header_map(ws)

    rows = []

    for row_idx in range(2, ws.max_row + 1):
        item = {"_ExcelOrder": row_idx - 2}
        empty = True

        for name, col in headers.items():
            value = ws.cell(row=row_idx, column=col).value
            if value is not None and value != "":
                empty = False
            item[name] = "" if value is None else value

        if not empty:
            rows.append(item)

    df = pd.DataFrame(rows)

    if df.empty:
        raise ValueError("Il foglio Carte è vuoto.")

    if "Quantità" not in df.columns:
        df["Quantità"] = 0

    df["Quantità"] = pd.to_numeric(df["Quantità"], errors="coerce").fillna(0).astype(int)

    if "Numero" in df.columns:
        df["Numero"] = df["Numero"].apply(format_number_text)

    print(f"Righe lette dall'Excel esistente: {len(df)}")

    return df


def clean_key_part(value):
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() == "nan":
        return ""

    return text


def normalize_variant_for_key(value):
    text = clean_key_part(value).lower()

    # Le righe JP manuali possono avere suffisso variabile. Lo rendo stabile.
    text = text.replace(" - jp manuale", "")
    text = text.replace("jp manuale", "")
    text = re.sub(r"\s+", " ", text).strip()

    return text


def normalize_idproduct_for_key(value):
    text = clean_key_part(value)

    if text == "":
        return ""

    try:
        return str(int(float(text)))
    except Exception:
        return text


def exact_key(row):
    return (
        normalize_code(clean_key_part(row.get("ID Carta", ""))),
        normalize_language(clean_key_part(row.get("Lingua", "EN"))),
        normalize_variant_for_key(row.get("Variante", "")),
        normalize_idproduct_for_key(row.get("Cardmarket idProduct", "")),
    )


def variant_key(row):
    return (
        normalize_code(clean_key_part(row.get("ID Carta", ""))),
        normalize_language(clean_key_part(row.get("Lingua", "EN"))),
        normalize_variant_for_key(row.get("Variante", "")),
    )


def card_lang_key(row):
    return (
        normalize_code(clean_key_part(row.get("ID Carta", ""))),
        normalize_language(clean_key_part(row.get("Lingua", "EN"))),
    )


def build_quantity_maps(existing_df):
    exact_qty = {}
    variant_qty = {}
    card_lang_qty = {}
    exact_order = {}
    variant_order = {}
    card_lang_orders = {}

    for _, row in existing_df.iterrows():
        qty = int(row.get("Quantità", 0) or 0)
        order = int(row.get("_ExcelOrder", 999999999))

        ek = exact_key(row)
        vk = variant_key(row)
        ck = card_lang_key(row)

        if ek not in exact_qty:
            exact_qty[ek] = qty
            exact_order[ek] = order

        if vk not in variant_qty:
            variant_qty[vk] = qty
            variant_order[vk] = order

        if ck not in card_lang_qty:
            card_lang_qty[ck] = []
            card_lang_orders[ck] = []

        card_lang_qty[ck].append(qty)
        card_lang_orders[ck].append(order)

    unique_card_lang_qty = {}
    unique_card_lang_order = {}

    for ck, values in card_lang_qty.items():
        # Fallback solo se per quella carta/lingua esiste una sola riga nel vecchio Excel.
        if len(values) == 1:
            unique_card_lang_qty[ck] = values[0]
            unique_card_lang_order[ck] = card_lang_orders[ck][0]

    return {
        "exact_qty": exact_qty,
        "variant_qty": variant_qty,
        "card_lang_qty": unique_card_lang_qty,
        "exact_order": exact_order,
        "variant_order": variant_order,
        "card_lang_order": unique_card_lang_order,
    }


def apply_existing_quantities(new_df, maps):
    new_df = new_df.copy()

    statuses = []
    quantities = []
    orders = []

    for _, row in new_df.iterrows():
        ek = exact_key(row)
        vk = variant_key(row)
        ck = card_lang_key(row)

        if ek in maps["exact_qty"]:
            quantities.append(maps["exact_qty"][ek])
            orders.append(maps["exact_order"].get(ek, 999999999))
            statuses.append("EXISTING_EXACT")
            continue

        if USE_FALLBACK_QUANTITY_MATCH and vk in maps["variant_qty"]:
            quantities.append(maps["variant_qty"][vk])
            orders.append(maps["variant_order"].get(vk, 999999999))
            statuses.append("EXISTING_VARIANT_FALLBACK")
            continue

        if USE_FALLBACK_QUANTITY_MATCH and ck in maps["card_lang_qty"]:
            quantities.append(maps["card_lang_qty"][ck])
            orders.append(maps["card_lang_order"].get(ck, 999999999))
            statuses.append("EXISTING_CARD_LANG_FALLBACK")
            continue

        quantities.append(0)
        orders.append(999999999)
        statuses.append("NEW_ROW_APPENDED")

    new_df["Quantità"] = quantities
    new_df["_SyncStatus"] = statuses
    new_df["_OldExcelOrder"] = orders

    return new_df


def scrape_bandai_fresh():
    sets = discover_sets()
    rows = []
    errors = []

    for i, s in enumerate(sets, start=1):
        series_id = str(s["series_id"])
        set_code = s.get("set_code", "")
        set_label = s.get("set_label", "")

        print(f"Scarico set {i}/{len(sets)}: {series_id} {set_code} {set_label}")

        try:
            response = safe_get(BASE_URL, params={"series": series_id})
            lines = get_lines_from_html(response.text)
            expected = parse_result_count(lines)
            cards = parse_cards_from_lines(lines, series_id)

            for card in cards:
                card["Set Code Ricerca"] = set_code
                card["Set Label Ricerca"] = set_label

                if card.get("Espansione") == "DON" and set_code:
                    card["Espansione"] = set_code
                    card["Numero"] = ""

            print(f"  Risultati dichiarati dal sito: {expected}")
            print(f"  Carte lette dallo script: {len(cards)}")

            if expected is not None and len(cards) < expected:
                print("  ATTENZIONE: carte lette meno dei risultati dichiarati.")

            rows.extend(cards)

        except Exception as e:
            print(f"Errore sul set {series_id}: {e}")
            errors.append({
                "Series ID": series_id,
                "Set Code": set_code,
                "Set Label": set_label,
                "Errore": str(e),
            })

        time.sleep(random.uniform(2.0, 5.0))

    raw_df = pd.DataFrame(rows)

    if raw_df.empty:
        raise ValueError("Nessuna carta scaricata dal sito Bandai.")

    raw_df = raw_df[
        raw_df["ID Carta"].notna()
        & (raw_df["ID Carta"].astype(str).str.strip() != "")
    ].copy()

    if "Numero" in raw_df.columns:
        raw_df["Numero"] = raw_df["Numero"].apply(format_number_text)

    dedupe_cols = [
        c for c in [
            "ID Carta",
            "Espansione",
            "Numero",
            "Rarità",
            "Tipo carta",
            "Nome",
            "Color",
            "Card Set(s)",
        ]
        if c in raw_df.columns
    ]

    if dedupe_cols:
        raw_df = raw_df.drop_duplicates(subset=dedupe_cols)

    raw_df.to_csv(SYNC_RAW_CSV, index=False, encoding="utf-8-sig")

    if errors:
        pd.DataFrame(errors).to_csv(ERRORS_CSV, index=False, encoding="utf-8-sig")

    print(f"Raw fresco salvato in: {SYNC_RAW_CSV}")

    return raw_df


def sort_existing_then_new(df):
    if not KEEP_EXISTING_ORDER_AND_APPEND_NEW:
        return df

    df = df.copy()
    df["_NewSort"] = df["_SyncStatus"].apply(lambda x: 1 if x == "NEW_ROW_APPENDED" else 0)

    # Nuove carte in fondo, in ordine naturale: espansione, numero, lingua, variante.
    sort_cols = ["_NewSort", "_OldExcelOrder", "Espansione", "Numero", "ID Carta", "Lingua", "Variante"]
    sort_cols = [c for c in sort_cols if c in df.columns]

    df = df.sort_values(sort_cols, kind="stable").copy()
    df = df.drop(columns=["_NewSort"])

    return df


def write_sync_report(df):
    report = df[[
        "ID Carta",
        "Espansione",
        "Numero",
        "Lingua",
        "Variante",
        "Quantità",
        "_SyncStatus",
        "_OldExcelOrder",
    ]].copy()

    report.to_csv(SYNC_REPORT_CSV, index=False, encoding="utf-8-sig")

    counts = report["_SyncStatus"].value_counts(dropna=False).to_dict()

    print("Report sync:")
    for key, value in counts.items():
        print(f"  {key}: {value}")

    print(f"Report salvato in: {SYNC_REPORT_CSV}")


def main():
    start_run_logging("sync")
    ensure_dirs()
    warn_out_extra_files()

    if BACKUP_FINAL_FILES:
        backup_file(OUTPUT_XLSX, move=False)
        backup_file(OUTPUT_JSON, move=False)

    existing_df = read_existing_excel()
    quantity_maps = build_quantity_maps(existing_df)

    df_cm = load_cardmarket_prices(CARDMARKET_MERGED_CSV)
    raw_df = scrape_bandai_fresh()

    fresh_df = enrich(raw_df, df_cm)
    fresh_df = fresh_df[fresh_df["Lingua"].isin(["EN", "JP"])].copy()

    if "Numero" in fresh_df.columns:
        fresh_df["Numero"] = fresh_df["Numero"].apply(format_number_text)

    if EXCLUDE_PROMOS:
        fresh_df = fresh_df[
            ~fresh_df["ID Carta"].astype(str).str.upper().str.startswith("P-")
        ].copy()

        if "CM Product Type" in fresh_df.columns:
            fresh_df = fresh_df[
                fresh_df["CM Product Type"].fillna("Standard") != "Promo"
            ].copy()

    dedupe_final = [
        c for c in [
            "ID Carta",
            "Lingua",
            "Variante",
            "Cardmarket idProduct",
            "Espansione",
            "Numero",
            "Rarità",
            "Tipo carta",
            "Nome",
            "Color",
            "Card Set(s)",
        ]
        if c in fresh_df.columns
    ]

    if dedupe_final:
        fresh_df = fresh_df.drop_duplicates(subset=dedupe_final)

    synced_df = apply_existing_quantities(fresh_df, quantity_maps)
    synced_df = sort_existing_then_new(synced_df)

    write_sync_report(synced_df)

    internal_cols = ["_SyncStatus", "_OldExcelOrder"]
    output_df = synced_df.drop(columns=[c for c in internal_cols if c in synced_df.columns]).copy()
    output_df = add_price_trends_from_latest_backup(output_df)
    save_price_trend_reports(output_df)
    output_df = sanitize_dataframe_for_excel(output_df)

    output_df.to_csv(SYNC_FINAL_CSV, index=False, encoding="utf-8-sig")
    print(f"CSV sync intermedio: {SYNC_FINAL_CSV}")

    build_excel(output_df)
    save_final_json_from_df(output_df, "one_piece_collection_sync.py")

    print("")
    print("Sync completato.")
    print("Ha riletto sito Bandai + JSON, mantenendo le Quantità dal file esistente.")
    print("File finali in out/:")
    print(f"- {OUTPUT_XLSX}")
    print(f"- {OUTPUT_JSON}")
    print("Intermedi in stg/:")
    print(f"- {SYNC_RAW_CSV}")
    print(f"- {SYNC_FINAL_CSV}")
    print(f"- {SYNC_REPORT_CSV}")


if __name__ == "__main__":
    main()
