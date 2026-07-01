import os
import re
import time
import random

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from one_piece_common import *

BASE_URL = "https://en.onepiece-cardgame.com/cardlist/"
JP_BASE_URL = "https://www.onepiece-cardgame.com/cardlist/"
CREATE_JP_ROWS = True
BACKUP_OLD_OUTPUTS = True
USE_EXISTING_RAW = True
CREATE_CHARTS = True

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/126.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9,it;q=0.8",
    "Connection": "close"
}


def clean(text):
    return " ".join(str(text).replace("\xa0", " ").split())


def clean_for_excel(value):
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    return value[:32000] if len(value) > 32000 else value


def sanitize_dataframe_for_excel(df):
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == "object":
            df[col] = df[col].apply(clean_for_excel)
    return df


def safe_get(url, params=None, max_retries=6, timeout=40):
    for attempt in range(1, max_retries + 1):
        try:
            r = requests.get(url, params=params, headers=HEADERS, timeout=timeout)
            r.raise_for_status()
            return r
        except requests.exceptions.RequestException as e:
            wait = 5 * attempt + random.uniform(1, 5)
            print(f"Errore richiesta tentativo {attempt}/{max_retries}: {e}")
            print(f"Attendo {wait:.1f} secondi...")
            time.sleep(wait)
    raise Exception(f"Richiesta fallita dopo {max_retries} tentativi: {url}")


def get_lines_from_html(html):
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n")
    return [clean(x) for x in text.split("\n") if clean(x)]


def discover_sets():
    print("Cerco gli ID dei set dal sito ufficiale...")
    r = safe_get(BASE_URL)
    soup = BeautifulSoup(r.text, "html.parser")
    discovered = {}
    for tag in soup.find_all(["a", "option", "input", "button", "li", "label", "span", "div"]):
        tag_text = clean(tag.get_text(" ", strip=True))
        vals = []
        for attr in ["href", "value", "data-value", "data-series", "data-series-id", "data-search", "data-id"]:
            v = tag.get(attr)
            if v:
                vals.append(str(v))
        joined = " ".join(vals + [tag_text])
        for series_id in re.findall(r"(?:series=|series%5B%5D=|series\[\]=)?(\d{6})", joined):
            cm = re.search(r"\[(OP-?\d{2}|ST-?\d{2}|EB-?\d{2}|PRB-?\d{2}|OP\d{2}-EB\d{2})\]", tag_text)
            code = cm.group(1).replace("-", "") if cm else ""
            discovered[series_id] = {"series_id": series_id, "set_code": code, "set_label": tag_text}
    for n in range(1, 31):
        discovered.setdefault(f"5690{n:02d}", {"series_id": f"5690{n:02d}", "set_code": f"ST{n:02d}", "set_label": f"ST-{n:02d}"})
    for n in range(1, 17):
        discovered.setdefault(f"5691{n:02d}", {"series_id": f"5691{n:02d}", "set_code": f"OP{n:02d}", "set_label": f"OP-{n:02d}"})
    for n in range(1, 4):
        discovered.setdefault(f"5692{n:02d}", {"series_id": f"5692{n:02d}", "set_code": f"EB{n:02d}", "set_label": f"EB-{n:02d}"})
    for n in range(1, 3):
        discovered.setdefault(f"5693{n:02d}", {"series_id": f"5693{n:02d}", "set_code": f"PRB{n:02d}", "set_label": f"PRB-{n:02d}"})
    discovered.setdefault("569801", {"series_id": "569801", "set_code": "", "set_label": "Other Product Card"})
    discovered.setdefault("569901", {"series_id": "569901", "set_code": "", "set_label": "Promotion card"})
    df = pd.DataFrame(list(discovered.values())).drop_duplicates(subset=["series_id"]).sort_values("series_id")
    df.to_csv(SETS_CSV, index=False, encoding="utf-8-sig")
    print(f"Set trovati/preparati: {len(df)}")
    return df.to_dict("records")


def parse_result_count(lines):
    for line in lines:
        m = re.search(r"(\d+)\s+results", line, re.I)
        if m:
            return int(m.group(1))
    return None


def extract_field(block_text, labels, label):
    label_pattern = "|".join(re.escape(x) for x in labels)
    pat = re.compile(rf"(?:###\s*)?{re.escape(label)}\s*\n+(.*?)(?=\n+(?:###\s*)?(?:{label_pattern})\s*\n+|\Z)", re.I | re.S)
    m = pat.search(block_text)
    if not m:
        return ""
    value = re.sub(r"[ \t]+", " ", re.sub(r"\n+", "\n", m.group(1).strip())).replace("TEXT VIEW", "").replace("CARD VIEW", "").replace("ボタン", "").strip()
    if label in ["Cost", "Life", "Attribute", "Power", "Counter", "Color", "Block", "Type"]:
        for line in value.splitlines():
            line = clean(line)
            if line and line.lower() != "icon":
                return line
        return ""
    return "\n".join(clean(x) for x in value.splitlines() if clean(x) and clean(x).lower() != "icon").strip()


def parse_don_cards(lines, series_id, normal_spans):
    text = "\n".join(lines)
    labels = ["Cost", "Life", "Attribute", "Power", "Counter", "Color", "Block", "Type", "Effect", "Trigger", "Card Set(s)", "Notes"]
    matches = list(re.finditer(r"(?:^|\n)(DON!! CARD|DON!!)\s*(?=\n)", text, re.I))
    cards, count = [], 0
    starts = [s for s, _ in normal_spans]
    for m in matches:
        if m.group(1).upper() != "DON!! CARD":
            continue
        start = m.start()
        if any(s <= start <= e for s, e in normal_spans):
            continue
        ends = [x for x in starts if x > start] + [x.start() for x in matches if x.start() > start and x.group(1).upper() == "DON!! CARD"]
        end = min(ends) if ends else len(text)
        block = text[start:end]
        if "Effect" not in block and "Card Set(s)" not in block and "Notes" not in block:
            continue
        count += 1
        cards.append({
            "Series ID": series_id, "ID Carta": f"DON-{series_id}-{count:03d}", "Espansione": "DON", "Numero": "",
            "Rarità": "DON!!", "Tipo carta": "DON!! CARD", "Nome": "DON!!", "Cost": "", "Life": "", "Attribute": "",
            "Power": "", "Counter": "", "Color": "", "Block": extract_field(block, labels, "Block"),
            "Type": extract_field(block, labels, "Type"), "Effect": extract_field(block, labels, "Effect"),
            "Trigger": extract_field(block, labels, "Trigger"), "Card Set(s)": extract_field(block, labels, "Card Set(s)"),
            "Notes": extract_field(block, labels, "Notes"), "Quantità": 0
        })
    if cards:
        print(f"  DON!! senza numero trovate: {len(cards)}")
    return cards


def parse_cards_from_lines(lines, series_id):
    text = "\n".join(lines)
    header_re = re.compile(r"([A-Z]{1,5}-?\d{2}-\d{3}|P-?\d{3})\s*\|\s*([^|\n]+?)\s*\|\s*(LEADER|CHARACTER|EVENT|STAGE|DON!! CARD|DON!!)", re.I)
    matches = list(header_re.finditer(text))
    print(f"  Header carta con codice trovati: {len(matches)}")
    labels = ["Cost", "Life", "Attribute", "Power", "Counter", "Color", "Block", "Type", "Effect", "Trigger", "Card Set(s)", "Notes"]
    cards, spans = [], []
    for pos, match in enumerate(matches):
        card_id = normalize_code(match.group(1).strip())
        rarity = match.group(2).strip()
        card_type = match.group(3).strip().upper()
        start = match.end()
        block_start = match.start()
        end = matches[pos + 1].start() if pos + 1 < len(matches) else len(text)
        spans.append((block_start, end))
        block = text[start:end]
        block_lines = [clean(x) for x in block.splitlines() if clean(x)]
        name = ""
        for line in block_lines:
            if line in ["TEXT VIEW", "CARD VIEW", "ボタン"] or line.startswith("Image: ") or line in labels or line.startswith("###") or line.lower() == "icon":
                continue
            name = line
            break
        exp, num = split_card_id(card_id)
        cards.append({
            "Series ID": series_id, "ID Carta": card_id, "Espansione": exp, "Numero": num, "Rarità": rarity,
            "Tipo carta": card_type, "Nome": name, "Cost": extract_field(block, labels, "Cost"),
            "Life": extract_field(block, labels, "Life"), "Attribute": extract_field(block, labels, "Attribute"),
            "Power": extract_field(block, labels, "Power"), "Counter": extract_field(block, labels, "Counter"),
            "Color": extract_field(block, labels, "Color"), "Block": extract_field(block, labels, "Block"),
            "Type": extract_field(block, labels, "Type"), "Effect": extract_field(block, labels, "Effect"),
            "Trigger": extract_field(block, labels, "Trigger"), "Card Set(s)": extract_field(block, labels, "Card Set(s)"),
            "Notes": extract_field(block, labels, "Notes"), "Quantità": 0
        })
    cards.extend(parse_don_cards(lines, series_id, spans))
    if not cards:
        print("  DEBUG: prime 80 righe ricevute dal sito:")
        for n, line in enumerate(lines[:80], start=1):
            print(f"    {n:03d}: {line}")
    return cards


def load_existing_raw():
    if not USE_EXISTING_RAW:
        return pd.DataFrame()
    if os.path.exists(RAW_CSV):
        print(f"Trovato raw CSV esistente: {RAW_CSV}")
        df = pd.read_csv(RAW_CSV, encoding="utf-8-sig", dtype={"Numero": str, "ID Carta": str})
        if "Numero" in df.columns:
            df["Numero"] = df["Numero"].apply(format_number_text)
        return df
    return pd.DataFrame()


def save_raw(rows, errors):
    if rows:
        df = pd.DataFrame(rows)
        if "Numero" in df.columns:
            df["Numero"] = df["Numero"].apply(format_number_text)
        dedupe = [c for c in ["ID Carta", "Espansione", "Numero", "Rarità", "Tipo carta", "Nome", "Color", "Card Set(s)"] if c in df.columns]
        if dedupe:
            df = df.drop_duplicates(subset=dedupe)
        df.to_csv(RAW_CSV, index=False, encoding="utf-8-sig")
    if errors:
        pd.DataFrame(errors).to_csv(ERRORS_CSV, index=False, encoding="utf-8-sig")



def scrape_jp_official_raw(sets, use_cache=True):
    """Scarica il catalogo ufficiale JP per ricavare rarità JP corrette.

    Il catalogo EN e quello JP possono differire nella rarità. Le carte JP della
    collezione usano questa tabella, più eventuali override locali.
    """
    if use_cache and os.path.exists(JP_OFFICIAL_RAW_CSV):
        print(f"Trovato raw JP esistente: {JP_OFFICIAL_RAW_CSV}")
        try:
            return pd.read_csv(JP_OFFICIAL_RAW_CSV, encoding="utf-8-sig", dtype={"Numero": str, "ID Carta": str})
        except Exception as exc:
            print(f"ATTENZIONE: cache JP non leggibile, la riscarico: {exc}")

    print("Scarico rarità ufficiali dal sito JP...")
    rows, errors = [], []
    for i, s in enumerate(sets, start=1):
        series_id = str(s["series_id"])
        set_code = s.get("set_code", "")
        set_label = s.get("set_label", "")
        print(f"Scarico JP set {i}/{len(sets)}: {series_id} {set_code} {set_label}")
        try:
            r = safe_get(JP_BASE_URL, params={"series": series_id})
            lines = get_lines_from_html(r.text)
            cards = parse_cards_from_lines(lines, series_id)
            for c in cards:
                c["Set Code Ricerca"] = set_code
                c["Set Label Ricerca"] = set_label
                c["Fonte catalogo"] = "JP ufficiale"
                if c.get("Espansione") == "DON" and set_code:
                    c["Espansione"] = set_code
                    c["Numero"] = ""
            print(f"  Carte JP lette: {len(cards)}")
            rows.extend(cards)
        except Exception as exc:
            print(f"Errore JP sul set {series_id}: {exc}")
            errors.append({"Series ID": series_id, "Set Code": set_code, "Set Label": set_label, "Errore": str(exc)})
        time.sleep(random.uniform(1.0, 3.0))

    jp_df = pd.DataFrame(rows)
    if not jp_df.empty:
        jp_df["ID Carta"] = jp_df["ID Carta"].apply(normalize_code)
        jp_df["Numero"] = jp_df["Numero"].apply(format_number_text)
        dedupe = [c for c in ["ID Carta", "Rarità", "Tipo carta", "Nome", "Color", "Card Set(s)"] if c in jp_df.columns]
        if dedupe:
            jp_df = jp_df.drop_duplicates(subset=dedupe)
        jp_df.to_csv(JP_OFFICIAL_RAW_CSV, index=False, encoding="utf-8-sig")
        print(f"Raw JP salvato in: {JP_OFFICIAL_RAW_CSV}")
    if errors:
        pd.DataFrame(errors).to_csv(os.path.join(STG_DIR, "one_piece_jp_errors.csv"), index=False, encoding="utf-8-sig")
    return jp_df


def ensure_default_jp_overrides_file():
    """Funzione mantenuta solo per compatibilità.

    Non crea più override: le rarità JP vengono ricavate dal sito ufficiale JP/Asia.
    """
    return None

def enrich(df_cards, df_cm):
    df_cards = df_cards.copy()
    if "Numero" in df_cards.columns:
        df_cards["Numero"] = df_cards["Numero"].apply(format_number_text)
    if EXCLUDE_PROMOS:
        before = len(df_cards)
        df_cards = df_cards[~df_cards["ID Carta"].astype(str).str.upper().str.startswith("P-")].copy()
        print(f"Promo P escluse dal database Bandai: {before - len(df_cards)}")
    for c in ["avg", "low", "trend", "avg1", "avg7", "avg30"]:
        if c in df_cm.columns:
            df_cm[c] = pd.to_numeric(df_cm[c], errors="coerce")
    price_col = PRICE_SOURCE_COLUMN if PRICE_SOURCE_COLUMN in df_cm.columns else "trend"
    merged = df_cards.merge(df_cm, on="ID Carta", how="left", suffixes=("", "_CM"))
    merged["Lingua"] = merged["CM Expansion Language"].apply(normalize_language)
    merged = merged[merged["Lingua"].isin(["EN", "JP"])].copy()
    merged["Variante"] = merged["Variante"].fillna("Base / non trovata su Cardmarket")
    merged["Cardmarket idProduct"] = merged["idProduct"]
    merged["Cardmarket Nome"] = merged["Cardmarket Nome"].fillna("")
    merged["Cardmarket Prodotti per carta"] = merged["Cardmarket Prodotti per carta"].fillna(0)
    merged["Valore"] = pd.to_numeric(merged[price_col], errors="coerce").fillna(0)
    def source(row):
        idp = row.get("idProduct", "")
        lang = row.get("Lingua", "")
        if pd.notna(idp) and str(idp).strip():
            try:
                return f"Cardmarket {price_col} per idProduct {int(idp)} [{lang}]"
            except Exception:
                return f"Cardmarket {price_col} per idProduct {idp} [{lang}]"
        return "Nessun prezzo Cardmarket"
    merged["Fonte prezzo"] = merged.apply(source, axis=1)
    merged["CM_Data_Prezzo"] = merged["Cardmarket Price Created At"].fillna("")
    merged["Valore totale"] = ""
    merged = merged.rename(columns={"low": "CM_Low", "trend": "CM_Trend", "avg": "CM_Avg", "avg1": "CM_Avg1", "avg7": "CM_Avg7", "avg30": "CM_Avg30"})
    if CREATE_JP_ROWS:
        existing_jp = set(merged.loc[merged["Lingua"] == "JP", "ID Carta"].dropna().astype(str))
        source_jp = merged[(merged["Lingua"] == "EN") & (~merged["ID Carta"].astype(str).isin(existing_jp))].copy()
        jp = source_jp.copy()
        jp["Lingua"] = "JP"
        jp["Valore"] = 0
        jp["Valore totale"] = ""
        jp["Fonte prezzo"] = "Manuale JP"
        jp["CM_Data_Prezzo"] = ""
        jp["Cardmarket idProduct"] = ""
        jp["Cardmarket Nome"] = ""
        jp["Cardmarket Prodotti per carta"] = ""
        jp["CM Expansion Name"] = ""
        jp["CM Expansion Code"] = ""
        jp["CM Expansion Language"] = "JP"
        jp["CM Expansion Product"] = ""
        jp["CM Product Type"] = "Standard"
        jp["Variante"] = jp["Variante"].apply(lambda x: f"{x} - JP manuale" if isinstance(x, str) else "JP manuale")
        merged = pd.concat([merged, jp], ignore_index=True)
    if EXCLUDE_PROMOS and "CM Product Type" in merged.columns:
        before = len(merged)
        merged = merged[merged["CM Product Type"].fillna("Standard") != "Promo"].copy()
        print(f"Promo escluse dal finale: {before - len(merged)}")
    merged["Numero"] = merged["Numero"].apply(format_number_text)
    merged = apply_jp_official_corrections(merged)
    return merged


def build_excel(df):
    create_collection_workbook_with_dashboard(df, OUTPUT_XLSX)


def main():
    start_run_logging("build")
    ensure_dirs()
    warn_out_extra_files()
    if BACKUP_OLD_OUTPUTS:
        backup_known_files(move=True)
    df_cm = load_cardmarket_prices(CARDMARKET_MERGED_CSV)
    sets = discover_sets()
    # Rileggo sempre il catalogo JP: serve per rarità diverse da EN, senza forzature manuali.
    scrape_jp_official_raw(sets, use_cache=False)
    existing = load_existing_raw()
    if not existing.empty and "Series ID" in existing.columns and "ID Carta" in existing.columns:
        existing = existing[existing["ID Carta"].notna() & (existing["ID Carta"].astype(str).str.strip() != "")]
        done = set(existing["Series ID"].astype(str).dropna())
        rows = existing.to_dict("records")
        print(f"Carte Bandai raw già presenti: {len(rows)}")
    else:
        done, rows = set(), []
    errors = []
    for i, s in enumerate(sets, start=1):
        series_id = str(s["series_id"])
        if series_id in done:
            print(f"Salto set già scaricato {i}/{len(sets)}: {series_id} {s.get('set_code','')} {s.get('set_label','')}")
            continue
        print(f"Scarico set {i}/{len(sets)}: {series_id} {s.get('set_code','')} {s.get('set_label','')}")
        try:
            r = safe_get(BASE_URL, params={"series": series_id})
            lines = get_lines_from_html(r.text)
            expected = parse_result_count(lines)
            cards = parse_cards_from_lines(lines, series_id)
            for c in cards:
                c["Set Code Ricerca"] = s.get("set_code", "")
                c["Set Label Ricerca"] = s.get("set_label", "")
                if c.get("Espansione") == "DON" and s.get("set_code"):
                    c["Espansione"] = s.get("set_code")
                    c["Numero"] = ""
            print(f"  Risultati dichiarati dal sito: {expected}")
            print(f"  Carte lette dallo script: {len(cards)}")
            rows.extend(cards)
            if cards:
                done.add(series_id)
            save_raw(rows, errors)
        except Exception as e:
            print(f"Errore sul set {series_id}: {e}")
            errors.append({"Series ID": series_id, "Set Code": s.get("set_code", ""), "Set Label": s.get("set_label", ""), "Errore": str(e)})
            save_raw(rows, errors)
        time.sleep(random.uniform(2.0, 5.0))
    raw_df = pd.DataFrame(rows)
    if raw_df.empty:
        print("Nessuna carta salvata. Non creo file finali.")
        return
    raw_df = raw_df[raw_df["ID Carta"].notna() & (raw_df["ID Carta"].astype(str).str.strip() != "")]
    raw_df["Numero"] = raw_df["Numero"].apply(format_number_text)
    dedupe = [c for c in ["ID Carta", "Espansione", "Numero", "Rarità", "Tipo carta", "Nome", "Color", "Card Set(s)"] if c in raw_df.columns]
    if dedupe:
        raw_df = raw_df.drop_duplicates(subset=dedupe)
    raw_df.to_csv(RAW_CSV, index=False, encoding="utf-8-sig")
    final_df = enrich(raw_df, df_cm)
    final_df = final_df[final_df["Lingua"].isin(["EN", "JP"])].copy()
    final_df["Numero"] = final_df["Numero"].apply(format_number_text)
    final_df = final_df[~final_df["ID Carta"].astype(str).str.upper().str.startswith("P-")].copy()
    if "CM Product Type" in final_df.columns:
        final_df = final_df[final_df["CM Product Type"].fillna("Standard") != "Promo"].copy()
    dedupe_final = [c for c in ["ID Carta", "Lingua", "Variante", "Cardmarket idProduct", "Espansione", "Numero", "Rarità", "Tipo carta", "Nome", "Color", "Card Set(s)"] if c in final_df.columns]
    if dedupe_final:
        final_df = final_df.drop_duplicates(subset=dedupe_final)
    final_df = apply_jp_official_corrections(final_df)
    final_df = add_price_trends_from_latest_backup(final_df)
    save_price_trend_reports(final_df)
    final_df = sanitize_dataframe_for_excel(final_df)
    final_df.to_csv(FINAL_STG_CSV, index=False, encoding="utf-8-sig")
    print(f"CSV finale intermedio: {FINAL_STG_CSV}")
    build_excel(final_df)
    save_final_json_from_df(final_df, "one_piece_collection_build.py")
    append_value_history(final_df, "build")
    print("\nFinito.")
    print("File finali in out/:")
    print(f"- {OUTPUT_XLSX}")
    print(f"- {OUTPUT_JSON}")
    print("Intermedi in stg/.")

if __name__ == "__main__":
    main()
