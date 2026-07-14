import os
import re
import pandas as pd
from openpyxl import load_workbook

from one_piece_common import *

KEEP_MANUAL_JP_IF_NO_PRICE = True


def variant_rank_from_text(text):
    if not isinstance(text, str):
        return None
    t = text.lower()
    if "base" in t:
        return 1
    m = re.search(r"(?:parallel|alt|reprint).*?(\d+)", t)
    if m:
        return int(m.group(1)) + 1
    m = re.search(r"variante\s+(\d+)", t)
    if m:
        return int(m.group(1))
    return None


def get_header_map(ws):
    return {str(ws.cell(row=1, column=c).value).strip(): c for c in range(1, ws.max_column + 1) if ws.cell(row=1, column=c).value}


def ensure_column(ws, headers, name):
    if name in headers:
        return headers[name]
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = name
    headers[name] = col
    return col


def get_cell(ws, row, headers, name):
    col = headers.get(name)
    if not col:
        return ""
    v = ws.cell(row=row, column=col).value
    return "" if v is None else v


def set_cell(ws, row, headers, name, value):
    col = ensure_column(ws, headers, name)
    ws.cell(row=row, column=col).value = value


def safe_int(value, default=0):
    """Converte in int senza esplodere su stringhe vuote/NaN.

    Alcune righe raw idProduct create dal fallback Cardmarket possono avere
    campi diagnostici vuoti, per esempio 'Cardmarket Prodotti per carta'.
    In quel caso non deve fermarsi tutto l'aggiornamento.
    """
    try:
        if value is None:
            return default
        if pd.isna(value):
            return default
    except Exception:
        pass
    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none", "null"}:
        return default
    try:
        return int(float(text.replace(",", ".")))
    except Exception:
        return default


def _norm_header_name(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).lower())


def canonicalize_cardmarket_headers(ws, headers):
    """Rende robusto il nome colonna Cardmarket idProduct.

    Nelle versioni precedenti possono essere rimaste varianti come
    'Cardmarket IDProduct' o 'Cardmarket IdProduct'. Se non le riconosciamo,
    Aggiorna valori non parte dall'idProduct e rischia di cadere nel fallback
    per ID Carta, scegliendo il prezzo sbagliato.
    """
    aliases = {
        "cardmarketidproduct",
        "cardmarketidprodotto",
        "cmidproduct",
        "idproduct",
        "idprodotto",
    }
    canonical = "Cardmarket idProduct"
    if canonical in headers:
        return headers
    for header, col in list(headers.items()):
        if _norm_header_name(header) in aliases:
            ws.cell(row=1, column=col).value = canonical
            headers.pop(header, None)
            headers[canonical] = col
            print(f"Colonna Cardmarket idProduct riconosciuta come '{header}' e normalizzata.")
            return headers
    return headers


def _has_idproduct(value):
    text = str(value).strip()
    return text != "" and text.lower() not in ["nan", "none", "null"]


def find_match(row_data, df_cm, by_idproduct=None):
    """Trova i dati Cardmarket.

    Regola principale: se sulla riga esiste un Cardmarket idProduct, l'update
    deve partire da quello e basta. In quel caso NON deve fare fallback per
    ID Carta, perché carte con lo stesso codice ma varianti diverse possono
    avere prezzi completamente diversi.
    """
    card_id = normalize_code(str(row_data.get("ID Carta", "")).strip())
    lang = normalize_language(str(row_data.get("Lingua", "EN")).strip())
    existing_idproduct = str(row_data.get("Cardmarket idProduct", "")).strip()
    variant = str(row_data.get("Variante", "")).strip()
    if not card_id or card_id.startswith("DON-"):
        return None

    # Prima scelta obbligatoria: idProduct già presente nella riga.
    if _has_idproduct(existing_idproduct):
        idp = normalize_idproduct_value(existing_idproduct)
        if idp is None:
            return None
        if by_idproduct and idp in by_idproduct:
            out = dict(by_idproduct[idp])
            out["__match_method"] = "Cardmarket idProduct"
            return out
        exact_idp = df_cm[pd.to_numeric(df_cm.get("idProduct"), errors="coerce").fillna(-1).astype(int) == idp]
        if not exact_idp.empty:
            out = exact_idp.iloc[0].to_dict()
            out["__match_method"] = "Cardmarket idProduct"
            return out
        # Importante: se l'utente ha indicato un idProduct, non uso fallback.
        # Meglio non aggiornare la riga che prendere il prezzo di una variante diversa.
        return None

    # Fallback solo se la riga non ha ancora Cardmarket idProduct.
    candidates = df_cm[df_cm["ID Carta"].astype(str).map(normalize_code) == card_id].copy()
    if candidates.empty:
        return None

    if "CM Expansion Language" in candidates.columns:
        lang_candidates = candidates[candidates["CM Expansion Language"].astype(str).map(normalize_language) == lang].copy()
        if not lang_candidates.empty:
            candidates = lang_candidates

    rank = variant_rank_from_text(variant)
    if rank is not None and "Cardmarket Variante N" in candidates.columns:
        exact = candidates[candidates["Cardmarket Variante N"] == rank]
        if not exact.empty:
            out = exact.iloc[0].to_dict()
            out["__match_method"] = "ID Carta + variante"
            return out

    out = candidates.iloc[0].to_dict()
    out["__match_method"] = "ID Carta fallback"
    return out

def apply_number_formats(ws, headers):
    for name in ["Valore", "Valore totale", "CM_Low", "CM_Trend", "CM_Avg", "CM_Avg1", "CM_Avg7", "CM_Avg30"]:
        col = headers.get(name)
        if col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = '€ #,##0.00'
    numero_col = headers.get("Numero")
    if numero_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=numero_col).number_format = "@"
            ws.cell(row=row, column=numero_col).value = format_number_text(ws.cell(row=row, column=numero_col).value)
    qty_col = headers.get("Quantità")
    if qty_col:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=qty_col).number_format = "0"


def workbook_to_dataframe(wb):
    ws = wb["Carte"]
    headers = [str(ws.cell(row=1, column=c).value).strip() if ws.cell(row=1, column=c).value else f"Column {c}" for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        item = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(row=r, column=c).value
            if v not in [None, ""]:
                empty = False
            item[h] = "" if v is None else v
        if not empty:
            rows.append(item)
    return pd.DataFrame(rows)



def normalize_idproduct_value(value):
    """Restituisce idProduct come int oppure None.

    Gestisce celle Excel salvate come numero, testo, float-string tipo '754942.0'
    e valori vuoti. Serve a rendere davvero primaria la chiave Cardmarket idProduct.
    """
    text = str(value).strip()
    if text == "" or text.lower() in ["nan", "none", "null"]:
        return None
    try:
        return int(float(text))
    except Exception:
        digits = re.sub(r"\D+", "", text)
        return int(digits) if digits else None


def _extract_code_from_cm_name(name):
    if not isinstance(name, str):
        return ""
    matches = re.findall(r"\(([A-Z]{1,5}-?\d{2}-\d{3}|P-?\d{3})\)", name)
    return normalize_code(matches[-1]) if matches else ""


def _extract_clean_cm_name(name):
    if not isinstance(name, str):
        return ""
    return re.sub(r"\s*\(([A-Z]{1,5}-?\d{2}-\d{3}|P-?\d{3})\)\s*$", "", name).strip()


def build_raw_idproduct_lookup_from_json():
    """Lookup idProduct diretto dai JSON Cardmarket non filtrati.

    Serve per i casi come Edward.Newgate 754942: il prodotto esiste nel
    products_singles e nel price_guide, ma può essere escluso dal catalogo
    filtrato perché classificato come promo/speciale. Se l'utente ha indicato
    esplicitamente quell'idProduct, il programma deve comunque usarlo.
    """
    lookup = {}
    try:
        price_file = find_json(PRICE_GUIDE_PATTERN)
        products_file = find_json(PRODUCTS_SINGLES_PATTERN)
        expansion_map = load_expansion_map(CARDMARKET_EXPANSIONS_CSV)
        with open(price_file, "r", encoding="utf-8") as f:
            price_data = json.load(f)
        with open(products_file, "r", encoding="utf-8") as f:
            products_data = json.load(f)
    except Exception as e:
        print(f"ATTENZIONE: lookup raw idProduct non disponibile: {e}")
        return lookup

    price_created_at = price_data.get("createdAt", "")
    prices = {}
    for item in price_data.get("priceGuides", []):
        idp = normalize_idproduct_value(item.get("idProduct", ""))
        if idp is not None:
            prices[idp] = item

    for product in products_data.get("products", []):
        idp = normalize_idproduct_value(product.get("idProduct", ""))
        if idp is None:
            continue
        price = prices.get(idp, {})
        name = product.get("name", "")
        card_id = _extract_code_from_cm_name(name)
        exp, num = split_card_id(card_id)
        id_expansion = product.get("idExpansion", "")
        def exp_info(key):
            return expansion_map.get(id_expansion, {}).get(key, "")
        row = {
            "idProduct": idp,
            "idCategory": product.get("idCategory", price.get("idCategory", "")),
            "categoryName": product.get("categoryName", ""),
            "idExpansion": id_expansion,
            "idMetacard": product.get("idMetacard", ""),
            "Cardmarket Nome": _extract_clean_cm_name(name),
            "ID Carta": card_id,
            "Espansione": exp,
            "Numero": format_number_text(num),
            "avg": clean_numeric(price.get("avg", "")),
            "low": clean_numeric(price.get("low", "")),
            "trend": clean_numeric(price.get("trend", "")),
            "avg1": clean_numeric(price.get("avg1", "")),
            "avg7": clean_numeric(price.get("avg7", "")),
            "avg30": clean_numeric(price.get("avg30", "")),
            "dateAdded": product.get("dateAdded", ""),
            "Cardmarket Price Created At": price_created_at,
            "CM Expansion Name": exp_info("Cardmarket Expansion Name"),
            "CM Expansion Code": exp_info("Cardmarket Expansion Code"),
            "CM Expansion Language": normalize_language(exp_info("Cardmarket Expansion Language") or "EN"),
            "CM Expansion Product": exp_info("Cardmarket Expansion Product"),
            "CM Product Type": "Standard",
            "Cardmarket Variante N": "",
            "Cardmarket Prodotti per carta": "",
            "__raw_idproduct_lookup": True,
        }
        try:
            row["CM Product Type"] = classify_cardmarket_product(pd.Series(row))
        except Exception:
            pass
        lookup[idp] = row
    return lookup


def build_idproduct_lookup(df_cm):
    lookup = {}
    if df_cm is not None and not df_cm.empty and "idProduct" in df_cm.columns:
        for _, row in df_cm.iterrows():
            idp = normalize_idproduct_value(row.get("idProduct", ""))
            if idp is not None:
                lookup[idp] = row.to_dict()

    raw_lookup = build_raw_idproduct_lookup_from_json()
    added = 0
    for idp, row in raw_lookup.items():
        if idp not in lookup:
            lookup[idp] = row
            added += 1
    if added:
        print(f"Lookup idProduct raw: aggiunti {added} idProduct esclusi dal catalogo filtrato.")
    return lookup


def cm_price_value(match, price_col):
    value = match.get(price_col, 0)
    if pd.isna(value) or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def apply_market_match_to_dataframe_row(df, idx, match, price_col, lang):
    """Aggiorna solo dati prezzo/mercato su una riga DataFrame.

    Non tocca mai anagrafica/manuali: Nome, Rarità, Lingua, Variante, Tipo carta,
    Color, Quantità e Cardmarket idProduct già presente.
    """
    idp = normalize_idproduct_value(match.get("idProduct", ""))
    idp_text = idp if idp is not None else match.get("idProduct", "")
    df.at[idx, "Valore"] = cm_price_value(match, price_col)
    df.at[idx, "Fonte prezzo"] = f"Cardmarket {price_col} per idProduct {idp_text} [{lang}]"
    df.at[idx, "CM_Data_Prezzo"] = match.get("Cardmarket Price Created At", "")
    # Cardmarket idProduct NON viene riscritto qui. Se l'utente lo ha messo, resta identico.
    for target, source in [
        ("Cardmarket Nome", "Cardmarket Nome"),
        ("Cardmarket Prodotti per carta", "Cardmarket Prodotti per carta"),
        ("CM_Low", "low"),
        ("CM_Trend", "trend"),
        ("CM_Avg", "avg"),
        ("CM_Avg1", "avg1"),
        ("CM_Avg7", "avg7"),
        ("CM_Avg30", "avg30"),
        ("CM Expansion Name", "CM Expansion Name"),
        ("CM Expansion Code", "CM Expansion Code"),
        ("CM Expansion Language", "CM Expansion Language"),
        ("CM Expansion Product", "CM Expansion Product"),
        ("CM Product Type", "CM Product Type"),
    ]:
        if target not in df.columns:
            df[target] = ""
        value = match.get(source, "")
        if target.startswith("CM_") and target != "CM_Data_Prezzo":
            value = clean_numeric(value)
        df.at[idx, target] = value


def enforce_idproduct_market_values(final_df, df_cm, price_col):
    """Passata finale di sicurezza.

    Il bug visto su Edward.Newgate OP02-001 nasceva perché una riga con
    Cardmarket idProduct 754942 finiva aggiornata col prezzo della base 696357.
    Questa funzione rende impossibile quel caso: ogni riga con idProduct viene
    riallineata direttamente a quella riga del price guide, senza fallback su ID Carta.
    """
    df = final_df.copy()
    if "Cardmarket idProduct" not in df.columns:
        return df, pd.DataFrame()
    lookup = build_idproduct_lookup(df_cm)
    audit = []
    for idx, row in df.iterrows():
        current_idp = normalize_idproduct_value(row.get("Cardmarket idProduct", ""))
        if current_idp is None:
            continue
        old_value = row.get("Valore", "")
        old_source = row.get("Fonte prezzo", "")
        match = lookup.get(current_idp)
        if match is None:
            audit.append({
                "row_index": idx,
                "ID Carta": row.get("ID Carta", ""),
                "Nome": row.get("Nome", ""),
                "Lingua": row.get("Lingua", ""),
                "Variante": row.get("Variante", ""),
                "Cardmarket idProduct": current_idp,
                "Status": "IDPRODUCT_NOT_FOUND",
                "Old Value": old_value,
                "New Value": old_value,
                "Old Fonte prezzo": old_source,
                "New Fonte prezzo": old_source,
            })
            continue
        lang = normalize_language(str(row.get("Lingua", match.get("CM Expansion Language", ""))))
        new_value = cm_price_value(match, price_col)
        apply_market_match_to_dataframe_row(df, idx, match, price_col, lang)
        audit.append({
            "row_index": idx,
            "ID Carta": row.get("ID Carta", ""),
            "Nome": row.get("Nome", ""),
            "Lingua": row.get("Lingua", ""),
            "Variante": row.get("Variante", ""),
            "Cardmarket idProduct": current_idp,
            "Status": "FORCED_BY_IDPRODUCT",
            "Old Value": old_value,
            "New Value": new_value,
            "Old Fonte prezzo": old_source,
            "New Fonte prezzo": df.at[idx, "Fonte prezzo"],
        })
    return df, pd.DataFrame(audit)

def update_excel_values():
    start_run_logging("update_values")
    ensure_dirs()
    warn_out_extra_files()
    if not os.path.exists(OUTPUT_XLSX):
        raise FileNotFoundError(f"Excel non trovato: {OUTPUT_XLSX}")
    df_cm = load_cardmarket_prices(CARDMARKET_UPDATED_CSV)
    by_idproduct = build_idproduct_lookup(df_cm)
    backup_file(OUTPUT_XLSX, move=False)
    backup_file(OUTPUT_JSON, move=False)
    wb = load_workbook(OUTPUT_XLSX)
    if "Carte" not in wb.sheetnames:
        raise ValueError("Nel file Excel non trovo il foglio 'Carte'.")
    ws = wb["Carte"]
    headers = get_header_map(ws)
    headers = canonicalize_cardmarket_headers(ws, headers)
    required = ["ID Carta", "Lingua", "Variante", "Valore", "Fonte prezzo", "CM_Data_Prezzo", "Cardmarket idProduct", "Cardmarket Nome", "Cardmarket Prodotti per carta", "CM_Low", "CM_Trend", "CM_Avg", "CM_Avg1", "CM_Avg7", "CM_Avg30", "CM Expansion Name", "CM Expansion Code", "CM Expansion Language", "CM Expansion Product", "CM Product Type"]
    for c in required:
        ensure_column(ws, headers, c)
    price_col = PRICE_SOURCE_COLUMN if PRICE_SOURCE_COLUMN in df_cm.columns else "trend"
    report = []
    updated = not_found = skipped_don = skipped_manual_jp = 0
    for row_idx in range(2, ws.max_row + 1):
        row_data = {name: get_cell(ws, row_idx, headers, name) for name in headers}
        card_id = normalize_code(str(row_data.get("ID Carta", "")).strip())
        lang = normalize_language(str(row_data.get("Lingua", "EN")).strip())
        old_value = get_cell(ws, row_idx, headers, "Valore")
        old_idp = get_cell(ws, row_idx, headers, "Cardmarket idProduct")
        if not card_id:
            continue
        if card_id.startswith("DON-"):
            skipped_don += 1
            report.append({"Excel row": row_idx, "ID Carta": card_id, "Lingua": lang, "Status": "SKIPPED_DON", "Old Value": old_value, "New Value": old_value, "Old idProduct": old_idp, "New idProduct": ""})
            continue
        match = find_match(row_data, df_cm, by_idproduct)
        match_method = match.get("__match_method", "") if isinstance(match, dict) else ""
        if match is None:
            if lang == "JP" and KEEP_MANUAL_JP_IF_NO_PRICE:
                skipped_manual_jp += 1
                status = "SKIPPED_MANUAL_JP_NO_PRICE"
            else:
                not_found += 1
                status = "IDPRODUCT_NOT_FOUND" if _has_idproduct(old_idp) else "NOT_FOUND"
            report.append({"Excel row": row_idx, "ID Carta": card_id, "Lingua": lang, "Status": status, "Match Method": "Cardmarket idProduct" if _has_idproduct(old_idp) else "", "Old Value": old_value, "New Value": old_value, "Old idProduct": old_idp, "New idProduct": ""})
            continue
        new_value = match.get(price_col, 0)
        if pd.isna(new_value):
            new_value = 0
        idp = match.get("idProduct", "")
        try:
            idp_text = int(idp)
        except Exception:
            idp_text = idp
        set_cell(ws, row_idx, headers, "Valore", float(new_value))
        set_cell(ws, row_idx, headers, "Fonte prezzo", f"Cardmarket {price_col} per idProduct {idp_text} [{lang}]")
        set_cell(ws, row_idx, headers, "CM_Data_Prezzo", match.get("Cardmarket Price Created At", ""))
        # Non aggiorno Cardmarket idProduct: se presente è la chiave usata per recuperare i dati.
        # Se manca del tutto, lo valorizzo solo come aiuto iniziale per righe nuove/non mappate.
        if str(old_idp).strip() in ["", "None", "nan", "NaN"]:
            set_cell(ws, row_idx, headers, "Cardmarket idProduct", idp_text)
        set_cell(ws, row_idx, headers, "Cardmarket Nome", match.get("Cardmarket Nome", ""))
        set_cell(ws, row_idx, headers, "Cardmarket Prodotti per carta", safe_int(match.get("Cardmarket Prodotti per carta", 0)))
        set_cell(ws, row_idx, headers, "CM_Low", clean_numeric(match.get("low", "")))
        set_cell(ws, row_idx, headers, "CM_Trend", clean_numeric(match.get("trend", "")))
        set_cell(ws, row_idx, headers, "CM_Avg", clean_numeric(match.get("avg", "")))
        set_cell(ws, row_idx, headers, "CM_Avg1", clean_numeric(match.get("avg1", "")))
        set_cell(ws, row_idx, headers, "CM_Avg7", clean_numeric(match.get("avg7", "")))
        set_cell(ws, row_idx, headers, "CM_Avg30", clean_numeric(match.get("avg30", "")))
        set_cell(ws, row_idx, headers, "CM Expansion Name", match.get("CM Expansion Name", ""))
        set_cell(ws, row_idx, headers, "CM Expansion Code", match.get("CM Expansion Code", ""))
        set_cell(ws, row_idx, headers, "CM Expansion Language", match.get("CM Expansion Language", lang))
        set_cell(ws, row_idx, headers, "CM Expansion Product", match.get("CM Expansion Product", ""))
        set_cell(ws, row_idx, headers, "CM Product Type", match.get("CM Product Type", "Standard"))
        updated += 1
        saved_idp = get_cell(ws, row_idx, headers, "Cardmarket idProduct")
        report.append({"Excel row": row_idx, "ID Carta": card_id, "Lingua": lang, "Status": "UPDATED", "Match Method": match_method, "Old Value": old_value, "New Value": float(new_value), "Old idProduct": old_idp, "New idProduct": saved_idp})
    apply_number_formats(ws, headers)
    # Converto il foglio aggiornato in DataFrame, calcolo trend contro il JSON precedente
    # e ricreo l'Excel completo con Dashboard/KPI/grafici.
    final_df = workbook_to_dataframe(wb)
    # Passata finale di sicurezza: se Cardmarket idProduct è presente, riallineo i valori
    # esclusivamente a quell'idProduct. Questo evita fallback accidentali su ID Carta.
    final_df, idproduct_audit = enforce_idproduct_market_values(final_df, df_cm, price_col)
    if not idproduct_audit.empty:
        idproduct_audit.to_csv(os.path.join(STG_DIR, "idproduct_update_audit.csv"), index=False, encoding="utf-8-sig")
        corrections = idproduct_audit[idproduct_audit["Status"] == "FORCED_BY_IDPRODUCT"]
        print(f"Audit idProduct: {len(corrections)} righe riallineate usando Cardmarket idProduct.")
        print(f"Audit idProduct salvato: {os.path.join(STG_DIR, 'idproduct_update_audit.csv')}")
    # Aggiornamento valori conservativo: da qui in poi non correggo campi catalogo/manuali
    # come nome, rarità, lingua, variante o quantità. Aggiorno solo prezzi/mercato.
    # Cardmarket idProduct resta stabile: viene usato come chiave, non sovrascritto.
    final_df = add_price_trends_from_latest_backup(final_df)
    save_price_trend_reports(final_df)
    final_df.to_csv(UPDATED_STG_CSV, index=False, encoding="utf-8-sig")
    create_collection_workbook_with_dashboard(final_df, OUTPUT_XLSX)
    save_final_json_from_df(final_df, "one_piece_collection_update_values.py")
    append_value_history(final_df, "update_values")
    pd.DataFrame(report).to_csv(UPDATE_REPORT_CSV, index=False, encoding="utf-8-sig")
    print("\nAggiornamento completato.")
    print(f"Carte aggiornate: {updated}")
    print(f"Non trovate: {not_found}")
    print(f"JP manuali senza prezzo lasciate intatte: {skipped_manual_jp}")
    print(f"DON saltate: {skipped_don}")
    try:
        method_counts = pd.DataFrame(report).get("Match Method", pd.Series(dtype=str)).value_counts(dropna=False).to_dict()
        print(f"Metodo match: {method_counts}")
    except Exception:
        pass
    print("File finali in out/:")
    print(f"- {OUTPUT_XLSX}")
    print(f"- {OUTPUT_JSON}")
    print("Intermedi in stg/.")

if __name__ == "__main__":
    update_excel_values()
